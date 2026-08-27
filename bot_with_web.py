import os, sys, logging, asyncio, secrets, functools
from datetime import datetime, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo
from contextlib import contextmanager
import jwt as pyjwt
from concurrent.futures import ThreadPoolExecutor

import psycopg2, psycopg2.extras
from psycopg2.pool import ThreadedConnectionPool
from aiohttp import web
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes, PicklePersistence

# ----------------------------------------------------------------------
# Environment variables
# ----------------------------------------------------------------------
BOT_TOKEN = os.environ.get("BOT_TOKEN")
ADMIN_GROUP_ID_STR = os.environ.get("ADMIN_GROUP_ID")
ADMIN_IDS_STR = os.environ.get("ADMIN_IDS", "")
DATABASE_URL = os.environ.get("DATABASE_URL")
WEBHOOK_URL = os.environ.get("RENDER_EXTERNAL_URL")
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET_TOKEN")
if not WEBHOOK_SECRET:
    WEBHOOK_SECRET = secrets.token_urlsafe(32)
    logging.warning(
        "WEBHOOK_SECRET_TOKEN غير مُعرَّف في متغيرات البيئة. تم توليد قيمة عشوائية مؤقتة "
        "(ستتغير عند كل إعادة تشغيل). يُنصح بشدة بتعيين WEBHOOK_SECRET_TOKEN ثابت في الإعدادات."
    )

TOPIC_RECLAMATIONS = int(os.environ.get("TOPIC_RECLAMATIONS", "0"))
TOPIC_VALIDATION = int(os.environ.get("TOPIC_VALIDATION", "0"))
TOPIC_VIDANGE = int(os.environ.get("TOPIC_VIDANGE", "0"))
TOPIC_VEHICLE_MGMT = int(os.environ.get("TOPIC_VEHICLE_MGMT", "0"))
TOPIC_GENERAL = int(os.environ.get("TOPIC_GENERAL", "0"))
TOPIC_HISTORY = int(os.environ.get("TOPIC_HISTORY", "0"))

if not all([BOT_TOKEN, ADMIN_GROUP_ID_STR, DATABASE_URL, WEBHOOK_URL]):
    sys.exit("FATAL: BOT_TOKEN, ADMIN_GROUP_ID, DATABASE_URL and WEBHOOK_URL must be set.")
if not all([TOPIC_RECLAMATIONS, TOPIC_VALIDATION, TOPIC_VIDANGE, TOPIC_VEHICLE_MGMT, TOPIC_GENERAL, TOPIC_HISTORY]):
    sys.exit("FATAL: All TOPIC_* environment variables must be set.")

try:
    ADMIN_GROUP_ID = int(ADMIN_GROUP_ID_STR)
except ValueError:
    sys.exit("FATAL: ADMIN_GROUP_ID must be an integer.")

ADMIN_IDS = set()
if ADMIN_IDS_STR:
    for uid in ADMIN_IDS_STR.split(","):
        uid = uid.strip()
        if uid.isdigit():
            ADMIN_IDS.add(int(uid))

TZ = ZoneInfo("Africa/Algiers")

# مسار ملف حفظ الحالة (user_data / bot_data) بين إعادات التشغيل.
# ملاحظة مهمة: هذا يحمي من فقدان الحالة فقط طالما القرص الذي يقيم عليه هذا
# المسار مستمر (persistent). في Render، القرص الافتراضي مؤقت (ephemeral) ويُمسح
# عند كل إعادة نشر (deploy) أو انتقال بين خوادم — لذا يجب ربط "Persistent Disk"
# فعلي في Render وتوجيه PERSISTENCE_PATH إلى مسار داخله كي يبقى الملف فعلاً
# بين عمليات إعادة التشغيل. راجع دليل النشر أسفل الرد.
PERSISTENCE_PATH = os.environ.get("PERSISTENCE_PATH", "/tmp/bot_persistence.pkl")

# مفتاح توقيع رموز JWT الخاصة بتطبيق أندرويد الإداري.
# يُفضَّل بشدة تعيين APP_JWT_SECRET ثابتاً في متغيرات البيئة، وإلا فكل جلسات
# تسجيل الدخول الحالية في التطبيق تنقطع (يحتاج المستخدمون لتسجيل دخول جديد)
# عند كل إعادة تشغيل للخادم.
APP_JWT_SECRET = os.environ.get("APP_JWT_SECRET")
if not APP_JWT_SECRET:
    APP_JWT_SECRET = secrets.token_urlsafe(32)
    logging.warning(
        "APP_JWT_SECRET غير مُعرَّف في متغيرات البيئة. تم توليد قيمة عشوائية مؤقتة "
        "(كل جلسات تطبيق أندرويد ستنقطع عند إعادة التشغيل). يُنصح بتعيين APP_JWT_SECRET ثابت."
    )
JWT_EXPIRY_DAYS = 30

# ----------------------------------------------------------------------
# Connection pool
# ----------------------------------------------------------------------
pool = None
def get_conn():
    global pool
    if pool is None:
        pool = ThreadedConnectionPool(2, 4, DATABASE_URL, sslmode='require')
    return pool.getconn()

def put_conn(conn):
    global pool
    if pool:
        pool.putconn(conn)

@contextmanager
def db_connection():
    conn = get_conn()
    try:
        yield conn
    finally:
        put_conn(conn)

# ----------------------------------------------------------------------
# Executor for offloading blocking DB / Excel work off the event loop
# ----------------------------------------------------------------------
_db_executor = ThreadPoolExecutor(max_workers=8, thread_name_prefix="db-worker")

async def run_db(func, *args, **kwargs):
    """Runs a blocking (sync) function in a worker thread so the asyncio
    event loop (and therefore the webhook server) stays responsive while
    a DB query / Excel generation is in progress."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(_db_executor, functools.partial(func, *args, **kwargs))

# ----------------------------------------------------------------------
# Database initialization
# ----------------------------------------------------------------------
def init_db():
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS drivers (
                user_id BIGINT PRIMARY KEY,
                name TEXT,
                vehicle TEXT,
                state TEXT,
                approval_status TEXT DEFAULT 'pending'
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS vehicles (
                code TEXT PRIMARY KEY
            )
        ''')
        cur.execute("SELECT COUNT(*) FROM vehicles")
        if cur.fetchone()[0] == 0:
            for v in ["F01","F02","H01"] + [f"M{i:02d}" for i in range(1,32)] + ["LOGAN"]:
                cur.execute("INSERT INTO vehicles (code) VALUES (%s) ON CONFLICT DO NOTHING", (v,))
        cur.execute('''
            CREATE TABLE IF NOT EXISTS problems (
                id SERIAL PRIMARY KEY,
                user_id BIGINT,
                driver_name TEXT,
                vehicle TEXT,
                problem_text TEXT,
                media_type TEXT,
                date TEXT,
                status TEXT DEFAULT 'قيد الانتظار',
                ruglee TEXT DEFAULT 'غير مُصلح',
                comments TEXT DEFAULT '',
                validation_requester BIGINT DEFAULT 0,
                group_message_id BIGINT DEFAULT 0
            )
        ''')
        cur.execute("ALTER TABLE problems ADD COLUMN IF NOT EXISTS validation_requester BIGINT DEFAULT 0")
        cur.execute("ALTER TABLE problems ADD COLUMN IF NOT EXISTS group_message_id BIGINT DEFAULT 0")
        cur.execute('''
            CREATE TABLE IF NOT EXISTS allowed_users (
                user_id BIGINT PRIMARY KEY,
                status TEXT DEFAULT 'approved'
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS help_videos (
                id SERIAL PRIMARY KEY,
                file_id TEXT,
                description TEXT DEFAULT ''
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS km_readings (
                id SERIAL PRIMARY KEY,
                vehicle TEXT,
                km INTEGER,
                date TEXT,
                driver_name TEXT DEFAULT ''
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS vehicle_vidange (
                vehicle TEXT PRIMARY KEY,
                last_vidange_km INTEGER DEFAULT 0
            )
        ''')
        cur.execute("SELECT code FROM vehicles")
        for v in [r[0] for r in cur.fetchall()]:
            cur.execute("INSERT INTO vehicle_vidange (vehicle, last_vidange_km) VALUES (%s, 0) ON CONFLICT DO NOTHING", (v,))

        # ---- Android admin app: roles + login codes ----
        cur.execute('''
            CREATE TABLE IF NOT EXISTS admins (
                user_id BIGINT PRIMARY KEY,
                role TEXT NOT NULL DEFAULT 'admin',
                display_name TEXT DEFAULT '',
                created_at TEXT
            )
        ''')
        # كل من كان في ADMIN_IDS (متغير البيئة) يصبح super_admin تلقائياً
        for uid in ADMIN_IDS:
            cur.execute(
                "INSERT INTO admins (user_id, role, created_at) VALUES (%s, 'super_admin', %s) "
                "ON CONFLICT (user_id) DO NOTHING",
                (uid, datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S"))
            )
        cur.execute('''
            CREATE TABLE IF NOT EXISTS app_login_codes (
                user_id BIGINT PRIMARY KEY,
                code TEXT NOT NULL,
                expires_at TEXT NOT NULL
            )
        ''')
        conn.commit()

# ----------------------------------------------------------------------
# Cached vehicle status (fixed)
# ----------------------------------------------------------------------
vehicle_cache = {}
cache_dirty = set()

def invalidate_cache(vehicle: str):
    cache_dirty.add(vehicle)
    vehicle_cache.pop(vehicle, None)

def refresh_cache(vehicle: str):
    with db_connection() as conn:
        cur = conn.cursor()

        # شكاوى سائق حقيقية (غير الفيدانج) بانتظار المعالجة -> أحمر
        cur.execute(
            "SELECT COUNT(*) FROM problems WHERE vehicle=%s AND status='قيد الانتظار' "
            "AND ruglee != 'تم الإصلاح' AND (media_type IS NULL OR media_type != 'نظام')",
            (vehicle,)
        )
        pending_normal = cur.fetchone()[0]

        # شكاوى سائق حقيقية قيد التصليح -> برتقالي
        cur.execute(
            "SELECT COUNT(*) FROM problems WHERE vehicle=%s AND status='قيد التصليح' "
            "AND ruglee != 'تم الإصلاح' AND (media_type IS NULL OR media_type != 'نظام')",
            (vehicle,)
        )
        in_progress_normal = cur.fetchone()[0]

        # مشاكل الفيدانج التلقائية غير المُصلحة (بصرف النظر عن حالتها) -> أبيض
        cur.execute(
            "SELECT COUNT(*) FROM problems WHERE vehicle=%s AND media_type='نظام' AND ruglee != 'تم الإصلاح'",
            (vehicle,)
        )
        vidange_open = cur.fetchone()[0]

        cur.execute("SELECT km FROM km_readings WHERE vehicle=%s ORDER BY date DESC LIMIT 1", (vehicle,))
        row = cur.fetchone()
        last_km = row[0] if row else None
        last_vid = get_last_vidange_km_noconn(conn, vehicle)
        # شبكة أمان: إن تجاوز العداد حد الفيدانج (9000 كم) ولم تُنشأ مشكلة النظام
        # بعد لأي سبب، نعتبرها حالة فيدانج (أبيض) وليست شكوى حمراء
        if not vidange_open and last_km and last_vid > 0 and last_km >= last_vid + 9000:
            vidange_open = 1

        if pending_normal > 0:
            status = 'bad'        # 🔴 شكوى سائق حقيقية بانتظار المعالجة
        elif in_progress_normal > 0:
            status = 'en_cours'   # 🟠 شكوى سائق قيد التصليح (حتى لو كان هناك فيدانج مستحق)
        elif vidange_open > 0:
            status = 'vidange'    # ⚪ لا شكاوى حقيقية مفتوحة، فقط فيدانج مستحق
        else:
            status = 'good'       # 🟢 كل شيء سليم

        cur.execute("SELECT COUNT(*) FROM problems WHERE vehicle=%s AND ruglee != 'تم الإصلاح'", (vehicle,))
        open_count = cur.fetchone()[0]
        if last_km and last_vid > 0:
            remaining = (last_vid + 10000) - last_km
        else:
            remaining = None
        vehicle_cache[vehicle] = {
            "status": status,
            "open_count": open_count,
            "remaining_km": remaining
        }

def get_vehicle_cache_entry(vehicle: str) -> dict:
    if vehicle in cache_dirty:
        refresh_cache(vehicle)
        cache_dirty.discard(vehicle)
    if vehicle not in vehicle_cache:
        refresh_cache(vehicle)
    return vehicle_cache[vehicle]

def get_vehicle_status_cached(vehicle: str) -> str:
    if vehicle in cache_dirty:
        refresh_cache(vehicle)
        cache_dirty.discard(vehicle)
    if vehicle not in vehicle_cache:
        refresh_cache(vehicle)
    return vehicle_cache[vehicle]["status"]

def count_open_problems_cached(vehicle: str) -> int:
    if vehicle in cache_dirty:
        refresh_cache(vehicle)
        cache_dirty.discard(vehicle)
    if vehicle not in vehicle_cache:
        refresh_cache(vehicle)
    return vehicle_cache[vehicle]["open_count"]

def get_remaining_km_cached(vehicle: str) -> int | None:
    if vehicle in cache_dirty:
        refresh_cache(vehicle)
        cache_dirty.discard(vehicle)
    if vehicle not in vehicle_cache:
        refresh_cache(vehicle)
    return vehicle_cache[vehicle]["remaining_km"]

def get_last_vidange_km_noconn(conn, vehicle: str) -> int:
    cur = conn.cursor()
    cur.execute("SELECT last_vidange_km FROM vehicle_vidange WHERE vehicle=%s", (vehicle,))
    row = cur.fetchone()
    return row[0] if row else 0

# ----------------------------------------------------------------------
# Database functions (unchanged)
# ----------------------------------------------------------------------
def get_driver(user_id: int) -> dict | None:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT name, vehicle, state, approval_status FROM drivers WHERE user_id = %s", (user_id,))
        row = cur.fetchone()
        return dict(row) if row else None

def set_driver(user_id: int, name=None, vehicle=None, state=None, approval_status=None):
    with db_connection() as conn:
        cur = conn.cursor()
        driver = get_driver(user_id)
        if driver:
            if name is not None:
                cur.execute("UPDATE drivers SET name=%s WHERE user_id=%s", (name, user_id))
            if vehicle is not None:
                cur.execute("UPDATE drivers SET vehicle=%s WHERE user_id=%s", (vehicle, user_id))
            if state is not None:
                cur.execute("UPDATE drivers SET state=%s WHERE user_id=%s", (state, user_id))
            if approval_status is not None:
                cur.execute("UPDATE drivers SET approval_status=%s WHERE user_id=%s", (approval_status, user_id))
        else:
            cur.execute("INSERT INTO drivers (user_id, name, vehicle, state, approval_status) VALUES (%s,%s,%s,%s,%s)",
                        (user_id, name or "", vehicle or "", state or "name_entry", approval_status or "pending"))
        conn.commit()

def get_all_vehicles():
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT code FROM vehicles ORDER BY code")
        rows = cur.fetchall()
        return [r[0] for r in rows]

def add_vehicle(code: str):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("INSERT INTO vehicles (code) VALUES (%s) ON CONFLICT DO NOTHING", (code,))
        cur.execute("INSERT INTO vehicle_vidange (vehicle, last_vidange_km) VALUES (%s, 0) ON CONFLICT DO NOTHING", (code,))
        conn.commit()

def remove_vehicle(code: str):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM vehicles WHERE code=%s", (code,))
        conn.commit()
    invalidate_cache(code)

def add_problem(user_id: int, driver_name: str, vehicle: str, problem_text: str, media_type: str, group_msg_id: int = 0) -> int:
    with db_connection() as conn:
        cur = conn.cursor()
        date = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(
            "INSERT INTO problems (user_id, driver_name, vehicle, problem_text, media_type, date, group_message_id) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (user_id, driver_name, vehicle, problem_text, media_type, date, group_msg_id)
        )
        problem_id = cur.fetchone()[0]
        conn.commit()
    invalidate_cache(vehicle)
    return problem_id

def update_problem_status(problem_id: int, status=None, ruglee=None, validation_requester=None, group_message_id=None):
    with db_connection() as conn:
        cur = conn.cursor()
        if status is not None:
            cur.execute("UPDATE problems SET status=%s WHERE id=%s", (status, problem_id))
        if ruglee is not None:
            cur.execute("UPDATE problems SET ruglee=%s WHERE id=%s", (ruglee, problem_id))
        if validation_requester is not None:
            cur.execute("UPDATE problems SET validation_requester=%s WHERE id=%s", (validation_requester, problem_id))
        if group_message_id is not None:
            cur.execute("UPDATE problems SET group_message_id=%s WHERE id=%s", (group_message_id, problem_id))
        conn.commit()
    problem = get_problem(problem_id)
    if problem:
        invalidate_cache(problem["vehicle"])

def set_problem_comment(problem_id: int, comment: str):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE problems SET comments=%s WHERE id=%s", (comment, problem_id))
        conn.commit()

def delete_problem(problem_id: int):
    problem = get_problem(problem_id)
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM problems WHERE id=%s", (problem_id,))
        conn.commit()
    if problem:
        invalidate_cache(problem["vehicle"])

def get_problem(problem_id: int) -> dict | None:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM problems WHERE id=%s", (problem_id,))
        row = cur.fetchone()
        return dict(row) if row else None

def get_all_problems():
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM problems ORDER BY date DESC")
        rows = cur.fetchall()
        return [dict(r) for r in rows]

def get_vehicle_history(vehicle: str) -> dict:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM problems WHERE vehicle=%s ORDER BY date DESC", (vehicle,))
        problems = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT date, km FROM km_readings WHERE vehicle=%s ORDER BY date DESC LIMIT 5", (vehicle,))
        readings = cur.fetchall()
    return {"problems": problems, "readings": readings}

def get_driver_problems(user_id: int, status_filter: str = None) -> list:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if status_filter:
            cur.execute("SELECT * FROM problems WHERE user_id=%s AND status=%s ORDER BY date DESC", (user_id, status_filter))
        else:
            cur.execute("SELECT * FROM problems WHERE user_id=%s ORDER BY date DESC", (user_id,))
        rows = cur.fetchall()
        return [dict(r) for r in rows]

def add_km_reading(vehicle: str, km: int, driver_name: str = ""):
    with db_connection() as conn:
        cur = conn.cursor()
        date = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
        cur.execute("INSERT INTO km_readings (vehicle, km, date, driver_name) VALUES (%s,%s,%s,%s)", (vehicle, km, date, driver_name))
        conn.commit()
    invalidate_cache(vehicle)

def fix_latest_km_reading(vehicle: str, corrected_km: int) -> tuple[bool, int | None]:
    """يصحّح آخر قراءة عداد مسجَّلة لمركبة (بدل حذفها) — يُستخدم عند إدخال
    السائق لرقم خاطئ عن طريق الخطأ. يُعيد (نجح, القيمة القديمة قبل التصحيح)."""
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, km FROM km_readings WHERE vehicle=%s ORDER BY date DESC LIMIT 1", (vehicle,))
        row = cur.fetchone()
        if not row:
            return False, None
        reading_id, old_km = row
        cur.execute("UPDATE km_readings SET km=%s WHERE id=%s", (corrected_km, reading_id))
        conn.commit()
    invalidate_cache(vehicle)
    return True, old_km

def get_latest_km(vehicle: str) -> int | None:
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT km FROM km_readings WHERE vehicle=%s ORDER BY date DESC LIMIT 1", (vehicle,))
        row = cur.fetchone()
        return row[0] if row else None

def get_last_vidange_km(vehicle: str) -> int:
    with db_connection() as conn:
        return get_last_vidange_km_noconn(conn, vehicle)

def set_last_vidange_km(vehicle: str, km: int):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE vehicle_vidange SET last_vidange_km=%s WHERE vehicle=%s", (km, vehicle))
        conn.commit()
    invalidate_cache(vehicle)

def has_active_vidange(vehicle: str) -> bool:
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM problems WHERE vehicle=%s AND media_type='نظام' AND ruglee != 'تم الإصلاح'", (vehicle,))
        row = cur.fetchone()
        return row is not None

def add_help_video(file_id: str, description: str = ""):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("INSERT INTO help_videos (file_id, description) VALUES (%s,%s)", (file_id, description))
        conn.commit()

def get_all_help_videos():
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM help_videos ORDER BY id DESC")
        rows = cur.fetchall()
        return [dict(r) for r in rows]

def delete_help_video(video_id: int):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM help_videos WHERE id=%s", (video_id,))
        conn.commit()

def add_allowed_user(user_id: int, status: str = "approved"):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("INSERT INTO allowed_users (user_id, status) VALUES (%s,%s) ON CONFLICT (user_id) DO UPDATE SET status=%s", (user_id, status, status))
        conn.commit()

def is_allowed(user_id: int) -> bool:
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT status FROM allowed_users WHERE user_id=%s", (user_id,))
        row = cur.fetchone()
        return row is not None and row[0] == 'approved'

def get_all_drivers():
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT user_id, name, vehicle FROM drivers WHERE approval_status='approved'")
        rows = cur.fetchall()
        return [dict(r) for r in rows]

def remove_driver(user_id: int):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM drivers WHERE user_id=%s", (user_id,))
        cur.execute("DELETE FROM allowed_users WHERE user_id=%s", (user_id,))
        conn.commit()

# ----------------------------------------------------------------------
# Admin roles + Android app login (used by the /api/* REST endpoints)
# ----------------------------------------------------------------------
def get_admin(user_id: int) -> dict | None:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT user_id, role, display_name FROM admins WHERE user_id=%s", (user_id,))
        row = cur.fetchone()
        return dict(row) if row else None

def list_admins() -> list:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT user_id, role, display_name, created_at FROM admins ORDER BY created_at")
        return [dict(r) for r in cur.fetchall()]

def upsert_admin(user_id: int, role: str, display_name: str = ""):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO admins (user_id, role, display_name, created_at) VALUES (%s,%s,%s,%s) "
            "ON CONFLICT (user_id) DO UPDATE SET role=%s, display_name=CASE WHEN %s != '' THEN %s ELSE admins.display_name END",
            (user_id, role, display_name, datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S"), role, display_name, display_name)
        )
        conn.commit()

def remove_admin(user_id: int):
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM admins WHERE user_id=%s", (user_id,))
        conn.commit()

def create_login_code(user_id: int) -> str:
    code = f"{secrets.randbelow(1_000_000):06d}"
    expires = (datetime.now(TZ) + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO app_login_codes (user_id, code, expires_at) VALUES (%s,%s,%s) "
            "ON CONFLICT (user_id) DO UPDATE SET code=%s, expires_at=%s",
            (user_id, code, expires, code, expires)
        )
        conn.commit()
    return code

def verify_and_consume_login_code(user_id: int, code: str) -> bool:
    with db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT code, expires_at FROM app_login_codes WHERE user_id=%s", (user_id,))
        row = cur.fetchone()
        if not row:
            return False
        stored_code, expires_at = row
        valid = (stored_code == code) and (datetime.now(TZ) <= datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S").replace(tzinfo=TZ))
        if valid:
            cur.execute("DELETE FROM app_login_codes WHERE user_id=%s", (user_id,))
            conn.commit()
        return valid

# ----------------------------------------------------------------------
# Excel generation
# ----------------------------------------------------------------------
def generate_problems_excel() -> BytesIO:
    problems = get_all_problems()
    wb = Workbook()
    ws = wb.active
    ws.title = "المشاكل"
    headers = ["التاريخ", "السائق", "المركبة", "المشكلة", "نوع الوسائط", "الحالة", "تم الإصلاح", "تعليقات"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for row_idx, p in enumerate(problems, 2):
        if p["ruglee"] == "تم الإصلاح":
            row_fill = green_fill
        elif p["status"] == "قيد الانتظار":
            row_fill = red_fill
        elif p["status"] == "قيد التصليح":
            row_fill = orange_fill
        else:
            row_fill = None
        vals = [p["date"], p["driver_name"], p["vehicle"], p["problem_text"], p["media_type"] or "—",
                p["status"], p["ruglee"], p["comments"] or ""]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            if row_fill:
                cell.fill = row_fill
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2, 50)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def generate_vidange_excel(vehicle_code: str = None) -> BytesIO:
    vehicles = [vehicle_code] if vehicle_code else get_all_vehicles()
    wb = Workbook()
    wb.remove(wb.active)
    for v in vehicles:
        ws = wb.create_sheet(title=v)
        ws.append(["التاريخ", "السائق", "العداد الحالي (KM)", "آخر فيدانج (KM)", "المسافة المتبقية للفيدانج"])
        with db_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT date, km, driver_name FROM km_readings WHERE vehicle=%s ORDER BY date DESC", (v,))
            readings = cur.fetchall()
        last_km = get_last_vidange_km(v)
        for date_str, km, dname in readings:
            remaining = (last_km + 10000) - km if last_km > 0 else "—"
            ws.append([date_str, dname or "", km, last_km, remaining])
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2, 50)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ----------------------------------------------------------------------
# Keyboards & Helpers
# ----------------------------------------------------------------------
MAIN_KEYBOARD = ReplyKeyboardMarkup([
    [KeyboardButton("📝 تقديم شكوى"), KeyboardButton("✅ طلب التحقق من الإصلاح")],
    [KeyboardButton("🚗 إدخال عدد الكيلومترات"), KeyboardButton("⚙️ الإعدادات")]
], resize_keyboard=True)

ADMIN_STICKY_KEYBOARD = ReplyKeyboardMarkup([
    [KeyboardButton("/admin")]
], resize_keyboard=True, one_time_keyboard=False)

def vehicle_inline_keyboard(vehicles: list, prefix="selv_") -> InlineKeyboardMarkup:
    buttons = [InlineKeyboardButton(v, callback_data=f"{prefix}{v}") for v in vehicles]
    return InlineKeyboardMarkup([buttons[i:i+4] for i in range(0, len(buttons), 4)])

def status_emoji_cached(vehicle: str) -> str:
    s = get_vehicle_status_cached(vehicle)
    if s == 'bad': return "🔴"
    if s == 'en_cours': return "🟠"
    if s == 'vidange': return "⚪"
    return "🟢"

def dashboard_button_text(vehicle: str) -> str:
    emoji = status_emoji_cached(vehicle)
    cnt = count_open_problems_cached(vehicle)
    rem = get_remaining_km_cached(vehicle)
    line1 = f"{emoji} {vehicle} ({cnt})"
    line2 = f"   {rem} كم" if rem is not None else "   —"
    return f"{line1}\n{line2}"

def admin_main_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🚘 تعديل المركبات", callback_data="admin_vehicles")],
        [InlineKeyboardButton("👤 إدارة السائقين", callback_data="admin_drivers")],
        [InlineKeyboardButton("🛢️ إدارة الفيدانج", callback_data="admin_vidange_menu")],
        [InlineKeyboardButton("📊 لوحة القيادة", callback_data="admin_dash")],
        [InlineKeyboardButton("📋 تصدير Excel", callback_data="admin_export_menu")],
    ])

def admin_vehicles_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ إضافة مركبة", callback_data="admin_addveh"),
         InlineKeyboardButton("➖ حذف مركبة", callback_data="admin_remveh")],
        [InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])

def admin_drivers_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ قبول سائقين", callback_data="admin_approve_list")],
        [InlineKeyboardButton("❌ حذف سائق", callback_data="admin_remove_driver_list")],
        [InlineKeyboardButton("📋 قائمة السائقين", callback_data="admin_drivers_list")],
        [InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])

def admin_vidange_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⚙️ تعيين فيدانج", callback_data="admin_setvid")],
        [InlineKeyboardButton("🚨 فيدانج عاجل", callback_data="admin_urgentvid")],
        [InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])

def admin_export_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 تصدير المشاكل", callback_data="admin_export")],
        [InlineKeyboardButton("🛢️ تصدير الفيدانج", callback_data="admin_vid")],
        [InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])

TOPIC_ACTIONS = {
    TOPIC_GENERAL: [
        [InlineKeyboardButton("📊 لوحة القيادة", callback_data="admin_dash")],
        [InlineKeyboardButton("📋 تصدير المشاكل", callback_data="admin_export")],
        [InlineKeyboardButton("🛢️ تصدير الفيدانج", callback_data="admin_vid")],
    ],
    TOPIC_RECLAMATIONS: [
        [InlineKeyboardButton("📊 لوحة القيادة", callback_data="admin_dash")],
        [InlineKeyboardButton("📋 تصدير المشاكل", callback_data="admin_export")],
    ],
    TOPIC_VALIDATION: [
        [InlineKeyboardButton("📊 لوحة القيادة", callback_data="admin_dash")],
        [InlineKeyboardButton("📋 تصدير المشاكل", callback_data="admin_export")],
    ],
    TOPIC_VIDANGE: [
        [InlineKeyboardButton("🛢️ تصدير الفيدانج", callback_data="admin_vid")],
        [InlineKeyboardButton("🚨 طلب فيدانج عاجل", callback_data="admin_urgentvid")],
        [InlineKeyboardButton("📋 تصدير المشاكل", callback_data="admin_export")],
    ],
    TOPIC_VEHICLE_MGMT: [
        [InlineKeyboardButton("➕ إضافة مركبة", callback_data="admin_addveh")],
        [InlineKeyboardButton("➖ حذف مركبة", callback_data="admin_remveh")],
        [InlineKeyboardButton("🚘 قائمة المركبات", callback_data="admin_listveh")],
    ],
    TOPIC_HISTORY: [
        [InlineKeyboardButton("📊 لوحة القيادة", callback_data="admin_dash")],
    ],
}

def get_topic_keyboard(thread_id: int) -> InlineKeyboardMarkup | None:
    actions = TOPIC_ACTIONS.get(thread_id)
    if actions:
        return InlineKeyboardMarkup(actions)
    return None

TELEGRAM_MAX_MESSAGE_LEN = 4096

def safe_text(text: str, limit: int = TELEGRAM_MAX_MESSAGE_LEN - 60) -> str:
    """Truncates long messages so send/edit never fails with
    'Message is too long' (Telegram's hard limit is 4096 chars)."""
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n… (تم اختصار القائمة، النتائج كثيرة جداً لعرضها كاملة)"

def status_icon_and_text(problem: dict) -> str:
    if problem["ruglee"] == "تم الإصلاح":
        return "🟢 مُصلح"
    if problem["status"] == "قيد الانتظار":
        if problem.get("media_type") == "نظام":
            return "⚪ قيد الانتظار (فيدانج)"
        return "🔴 قيد الانتظار"
    if problem["status"] == "قيد التصليح":
        return "🟠 قيد التصليح"
    return "⚪ حالة غير معروفة"

def build_problem_keyboard(problem_id: int) -> InlineKeyboardMarkup:
    if problem_id == 0:
        return InlineKeyboardMarkup([])
    problem = get_problem(problem_id)
    if not problem:
        return InlineKeyboardMarkup([])
    status = problem["status"]
    ruglee = problem["ruglee"]
    row1 = []
    if status == "قيد الانتظار":
        row1.append(InlineKeyboardButton("🔧 قيد التصليح", callback_data=f"val_{problem_id}"))
    elif status == "قيد التصليح":
        row1.append(InlineKeyboardButton("⏳ إعادة للانتظار", callback_data=f"val_{problem_id}"))
        if problem["media_type"] and not problem["comments"]:
            row1.append(InlineKeyboardButton("✅ تم الإصلاح (تعليق مطلوب)", callback_data=f"fix_comment_{problem_id}"))
        else:
            row1.append(InlineKeyboardButton("✅ تم الإصلاح", callback_data=f"rug_{problem_id}"))
    rows = []
    if row1:
        rows.append(row1)
    rows.append([InlineKeyboardButton("💬 تعليق", callback_data=f"com_{problem_id}")])
    if ruglee != "تم الإصلاح":
        rows.append([InlineKeyboardButton("🗑️ حذف", callback_data=f"del_{problem_id}")])
    return InlineKeyboardMarkup(rows)

def update_status_line(problem: dict, new_status: str = None, new_ruglee: str = None) -> str:
    dname = problem["driver_name"]
    veh = problem["vehicle"]
    prob_text = problem["problem_text"]
    status_line = f"الحالة: {status_icon_and_text({'status': new_status or problem['status'], 'ruglee': new_ruglee or problem['ruglee'], 'media_type': problem.get('media_type')})}"
    return f"السائق: {dname}\nالمركبة: {veh}\nالمشكلة: {prob_text}\n{status_line}"

async def _update_problem_message(problem: dict, new_status: str = None, new_ruglee: str = None):
    """Synchronize the original reclamation message with its database state."""
    if not problem or not problem.get("group_message_id"):
        return
    new_text = update_status_line(problem, new_status=new_status, new_ruglee=new_ruglee)
    try:
        if problem["media_type"] and problem["media_type"] != "نظام":
            await app.bot.edit_message_caption(
                chat_id=ADMIN_GROUP_ID,
                message_id=problem["group_message_id"],
                caption=new_text
            )
        else:
            await app.bot.edit_message_text(
                chat_id=ADMIN_GROUP_ID,
                message_id=problem["group_message_id"],
                text=new_text
            )
    except Exception as e:
        logging.warning(f"Could not update problem message: {e}")

# ----------------------------------------------------------------------
# Cancel handler
# ----------------------------------------------------------------------
async def cancel_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data.clear()
    if update.effective_chat.type == "private":
        driver = get_driver(query.from_user.id)
        if driver and driver.get("approval_status") == "approved":
            set_driver(query.from_user.id, state="idle")
            await query.edit_message_text("تم الإلغاء.")
            await context.bot.send_message(chat_id=query.from_user.id,
                                           text="يمكنك الآن استخدام الأزرار أدناه:",
                                           reply_markup=MAIN_KEYBOARD)
            return
    await query.edit_message_text("تم الإلغاء.")

# ----------------------------------------------------------------------
# Core handlers
# ----------------------------------------------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in ADMIN_IDS:
        await run_db(add_allowed_user, user_id)
        await run_db(set_driver, user_id, approval_status="approved")
    driver = await run_db(get_driver, user_id)
    if driver and driver["name"] and driver["vehicle"] and driver.get("approval_status") == "approved":
        if driver["state"] != "idle":
            await run_db(set_driver, user_id, state="idle")
        await update.message.reply_text(
            f"أهلاً بعودتك، {driver['name']}!\nمركبتك: {driver['vehicle']}",
            reply_markup=MAIN_KEYBOARD
        )
    elif driver and driver.get("approval_status") == "pending":
        await update.message.reply_text("شكراً لتسجيلك. طلب صلاحيتك قيد المراجعة. انتظر قبول المشرف.")
    else:
        await run_db(set_driver, user_id, state="name_entry", approval_status="pending")
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]])
        await update.message.reply_text("مرحباً! الرجاء إدخال اسمك الكامل:", reply_markup=markup)

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return
    user_id = update.effective_user.id
    text = update.message.text.strip()
    driver = await run_db(get_driver, user_id)
    state = driver["state"] if driver else "name_entry"

    # Block non-allowed users
    if user_id not in ADMIN_IDS and not await run_db(is_allowed, user_id):
        if state == "name_entry":
            await run_db(set_driver, user_id, name=text, state="awaiting_approval", approval_status="pending")
            username = update.effective_user.username
            mention = f"@{username}" if username else f"[{text}](tg://user?id={user_id})"
            msg = await context.bot.send_message(
                chat_id=ADMIN_GROUP_ID,
                message_thread_id=TOPIC_VALIDATION,
                text=f"📌 طلب صلاحية جديدة:\nالاسم: {text}\nالمستخدم: {mention}\nالحالة: ⏳ قيد الانتظار",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ قبول", callback_data=f"approve_{user_id}"),
                     InlineKeyboardButton("❌ رفض", callback_data=f"reject_{user_id}")]
                ])
            )
            context.bot_data.setdefault("approval_msg", {})[user_id] = msg.message_id
            await update.message.reply_text("تم إرسال طلبك للمراجعة. ستصلك رسالة عند القبول.")
            return
        else:
            await update.message.reply_text("غير مصرح لك باستخدام البوت حالياً.")
            return

    # Comment session
    if context.user_data.get("awaiting_comment"):
        problem_id = context.user_data.pop("awaiting_comment")
        await run_db(set_problem_comment, problem_id, text)
        await update.message.reply_text("✅ تم حفظ التعليق بنجاح.", reply_markup=MAIN_KEYBOARD)
        return

    # KM after vidange repair
    if context.user_data.get("await_km"):
        vehicle = context.user_data["await_km_vehicle"]
        if text.isdigit():
            km = int(text)
            last_km = await run_db(get_latest_km, vehicle)
            if last_km is not None and km <= last_km:
                await update.message.reply_text(f"⚠️ الكيلومتر يجب أن يكون أكبر من آخر قراءة ({last_km} كم). أعد إدخال القيمة الصحيحة.",
                                                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
                return
            driver_name = driver["name"] if driver else "Unknown"
            msg = await context.bot.send_message(
                chat_id=ADMIN_GROUP_ID,
                message_thread_id=TOPIC_VIDANGE,
                text=f"⚙️ تأكيد تحديث الفيدانج:\nالمركبة: {vehicle}\nالقيمة المدخلة: {km} كم\nالسائق: {driver_name}\nالحالة: ⏳ في انتظار التأكيد",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ تأكيد", callback_data=f"vidconfirm_{user_id}_{km}"),
                     InlineKeyboardButton("✏️ تعديل", callback_data=f"vidmodify_{user_id}")]
                ])
            )
            context.bot_data.setdefault("pending_vidange", {})[msg.message_id] = {
                "user_id": user_id,
                "vehicle": vehicle,
                "km": km
            }
            await update.message.reply_text("تم إرسال القيمة للمراجعة من قبل المشرفين.", reply_markup=MAIN_KEYBOARD)
            context.user_data.pop("await_km_vehicle", None)
            context.user_data.clear()
            return
        else:
            await update.message.reply_text("الرجاء إرسال رقم صحيح.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
            return

    # State machine
    if state == "name_entry":
        # Validate name
        allowed = True
        for c in text:
            if not (c.isalnum() or c.isspace() or ('\u0600' <= c <= '\u06ff') or c in '-ء'):
                allowed = False
                break
        if not text or not allowed:
            await update.message.reply_text("⚠️ الاسم يجب أن يحتوي على أحرف وأرقام ومسافات فقط. أعد الإدخال.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
            return
        # If driver already has a vehicle, this is a name change, not initial registration
        if driver and driver.get("vehicle"):
            await run_db(set_driver, user_id, name=text, state="idle")
            await update.message.reply_text(f"تم تغيير الاسم إلى {text}.", reply_markup=MAIN_KEYBOARD)
            return
        # Otherwise, initial registration
        await run_db(set_driver, user_id, name=text, state="vehicle_selection")
        vehicles = await run_db(get_all_vehicles)
        await update.message.reply_text("تم حفظ الاسم. اختر مركبتك:", reply_markup=vehicle_inline_keyboard(vehicles, "selv_"))
        return

    if state == "vehicle_selection":
        vehicles = await run_db(get_all_vehicles)
        await update.message.reply_text("الرجاء اختيار المركبة من القائمة:", reply_markup=vehicle_inline_keyboard(vehicles, "selv_"))
        return

    # Main keyboard
    if text == "📝 تقديم شكوى":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]])
        await update.message.reply_text("أرسل وصف المشكلة (نص، صورة، فيديو، أو صوت).", reply_markup=markup)
        context.user_data["expecting_reclamation"] = True
        return
    if text == "✅ طلب التحقق من الإصلاح":
        problems = await run_db(get_driver_problems, user_id, "قيد التصليح")
        if not problems:
            await update.message.reply_text("لا توجد مشاكل بحاجة للتحقق من إصلاحها.", reply_markup=MAIN_KEYBOARD)
            return
        pending = [p for p in problems if p["validation_requester"] == 0]
        if not pending:
            await update.message.reply_text("جميع المشاكل قيد التحقق بالفعل.", reply_markup=MAIN_KEYBOARD)
            return
        buttons = [InlineKeyboardButton(f"مشكلة #{p['id']} - {p['problem_text'][:30]}...", callback_data=f"valreq_{p['id']}") for p in pending]
        await update.message.reply_text("اختر المشكلة التي تم إصلاحها:", reply_markup=InlineKeyboardMarkup([buttons[i:i+2] for i in range(0, len(buttons), 2)]))
        return
    if text == "🚗 إدخال عدد الكيلومترات":
        if not driver or not driver["vehicle"]:
            await update.message.reply_text("يجب إكمال الملف أولاً.")
            return
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]])
        await update.message.reply_text(f"أرسل عدد الكيلومترات الحالي للمركبة {driver['vehicle']}:", reply_markup=markup)
        context.user_data["expecting_km"] = True
        return
    if text == "⚙️ الإعدادات":
        markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("🚘 تغيير المركبة", callback_data="settings_change_veh")],
            [InlineKeyboardButton("📜 سجل شكاويي", callback_data="settings_history")],
            [InlineKeyboardButton("✏️ تعديل الاسم", callback_data="settings_change_name")],
            [InlineKeyboardButton("❓ مساعدة", callback_data="settings_help")]
        ])
        await update.message.reply_text("اختر الإعداد المطلوب:", reply_markup=markup)
        return

    # Reclamation / km input
    if context.user_data.get("expecting_reclamation"):
        if not text:
            await update.message.reply_text("الرجاء كتابة وصف للمشكلة. لا يمكن إرسال شكوى فارغة.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
            return
        context.user_data.pop("expecting_reclamation")
        if not driver or not driver["name"] or not driver["vehicle"]:
            await update.message.reply_text("ملفك غير مكتمل.")
            return
        status_line = "🔴 الحالة: قيد الانتظار"
        report = f"السائق: {driver['name']}\nالمركبة: {driver['vehicle']}\nالمشكلة: {text}\n{status_line}"
        msg = await context.bot.send_message(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_RECLAMATIONS, text=report,
                                             reply_markup=build_problem_keyboard(0))
        problem_id = await run_db(add_problem, user_id, driver["name"], driver["vehicle"], text, "", group_msg_id=msg.message_id)
        await msg.edit_reply_markup(reply_markup=await run_db(build_problem_keyboard, problem_id))
        await update.message.reply_text("تم إرسال الشكوى.", reply_markup=MAIN_KEYBOARD)
        return
    if context.user_data.get("expecting_km"):
        if not text.isdigit():
            await update.message.reply_text("يجب إرسال رقم.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
            return
        km = int(text)
        vehicle = driver["vehicle"] if driver else None
        if not vehicle:
            await update.message.reply_text("ملف غير مكتمل.")
            return
        last_km = await run_db(get_latest_km, vehicle)
        if last_km is not None and km <= last_km:
            await update.message.reply_text(f"⚠️ الكيلومتر يجب أن يكون أكبر من آخر قراءة ({last_km} كم). أعد إدخال القيمة الصحيحة.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
            return
        context.user_data.pop("expecting_km")
        await run_db(add_km_reading, vehicle, km, driver_name=driver["name"])
        last_vid = await run_db(get_last_vidange_km, vehicle)
        if last_vid > 0 and km >= last_vid + 9000 and not await run_db(has_active_vidange, vehicle):
            vidange_problem_id = await run_db(add_problem, user_id, f"{driver['name']} (نظام)", vehicle, f"Vidange {vehicle}", "نظام")
            await context.bot.send_message(
                chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_VIDANGE,
                text=f"⚠️ تنبيه فيدانج: المركبة {vehicle}\nالعداد الحالي: {km} كم\nآخر فيدانج: {last_vid} كم\n⚪ الحالة: قيد الانتظار",
                reply_markup=await run_db(build_problem_keyboard, vidange_problem_id)
            )
        await update.message.reply_text(f"تم تسجيل العداد: {km} كم.", reply_markup=MAIN_KEYBOARD)
        return

    await update.message.reply_text("استخدم الأزرار أدناه.", reply_markup=MAIN_KEYBOARD)

async def handle_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return
    user_id = update.effective_user.id
    driver = get_driver(user_id)
    if not driver or not driver["name"] or not driver["vehicle"] or driver.get("approval_status") != "approved":
        await update.message.reply_text("ملفك غير مكتمل.")
        return
    caption = update.message.caption or "(مرفق وسائط)"
    header = f"السائق: {driver['name']}\nالمركبة: {driver['vehicle']}\nالمشكلة: {caption}\n🔴 الحالة: قيد الانتظار"
    if update.message.photo:
        file_id = update.message.photo[-1].file_id
        media_type = "صورة"
        msg = await context.bot.send_photo(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_RECLAMATIONS, photo=file_id,
                                           caption=header, reply_markup=build_problem_keyboard(0))
        problem_id = add_problem(user_id, driver["name"], driver["vehicle"], caption, media_type, group_msg_id=msg.message_id)
        await msg.edit_reply_markup(reply_markup=build_problem_keyboard(problem_id))
    elif update.message.video:
        file_id = update.message.video.file_id
        media_type = "فيديو"
        msg = await context.bot.send_video(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_RECLAMATIONS, video=file_id,
                                           caption=header, reply_markup=build_problem_keyboard(0))
        problem_id = add_problem(user_id, driver["name"], driver["vehicle"], caption, media_type, group_msg_id=msg.message_id)
        await msg.edit_reply_markup(reply_markup=build_problem_keyboard(problem_id))
    elif update.message.voice:
        file_id = update.message.voice.file_id
        media_type = "صوت"
        msg = await context.bot.send_voice(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_RECLAMATIONS, voice=file_id,
                                           caption=header, reply_markup=build_problem_keyboard(0))
        problem_id = add_problem(user_id, driver["name"], driver["vehicle"], caption, media_type, group_msg_id=msg.message_id)
        await msg.edit_reply_markup(reply_markup=build_problem_keyboard(problem_id))
    await update.message.reply_text("تم إرسال الشكوى.", reply_markup=MAIN_KEYBOARD)

# ----------------------------------------------------------------------
# Callback handlers (modified)
# ----------------------------------------------------------------------
async def vehicle_selection_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    vehicle = query.data.split("_", 1)[1]
    user_id = query.from_user.id
    context.user_data["confirm_vehicle"] = vehicle
    await query.edit_message_text(f"هل تريد تعيين المركبة {vehicle}؟", reply_markup=InlineKeyboardMarkup([
        [InlineKeyboardButton("نعم", callback_data=f"confirmveh_{vehicle}"),
         InlineKeyboardButton("إلغاء", callback_data="cancel_veh")]
    ]))

async def confirm_vehicle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    vehicle = query.data.split("_", 1)[1]
    user_id = query.from_user.id
    set_driver(user_id, vehicle=vehicle, state="idle")
    await query.edit_message_text(f"تم تعيين المركبة إلى {vehicle}.")
    await context.bot.send_message(chat_id=user_id, text="يمكنك الآن استخدام الأزرار أدناه:", reply_markup=MAIN_KEYBOARD)

async def cancel_vehicle_selection_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    vehicles = get_all_vehicles()
    await query.edit_message_text("اختر مركبتك:", reply_markup=vehicle_inline_keyboard(vehicles, "selv_"))

async def valide_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    problem = get_problem(problem_id)
    if not problem: return await query.answer("المشكلة غير موجودة.")
    new_status = "قيد التصليح" if problem["status"] == "قيد الانتظار" else "قيد الانتظار"
    await run_db(update_problem_status, problem_id, status=new_status)
    updated_problem = await run_db(get_problem, problem_id)
    await _update_problem_message(updated_problem)
    await query.edit_message_reply_markup(reply_markup=await run_db(build_problem_keyboard, problem_id))

async def ruglee_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    problem = get_problem(problem_id)
    if not problem: return await query.answer("غير موجود.")
    if problem["media_type"] and not problem["comments"]:
        await query.answer("يجب إضافة تعليق أولاً قبل تأكيد الإصلاح.", show_alert=True)
        await context.bot.send_message(chat_id=query.from_user.id, text="يجب إضافة تعليق للمشكلة قبل وضعها كمُصلحة. أرسل التعليق هنا.")
        context.user_data["awaiting_comment"] = problem_id
        return
    new_ruglee = "تم الإصلاح" if problem["ruglee"] == "غير مُصلح" else "غير مُصلح"
    await run_db(update_problem_status, problem_id, ruglee=new_ruglee)
    updated_problem = await run_db(get_problem, problem_id)
    await _update_problem_message(updated_problem)
    val_msg_id = context.bot_data.get("validation_msgs", {}).pop(problem_id, None)
    if val_msg_id:
        try:
            await context.bot.edit_message_text(
                chat_id=ADMIN_GROUP_ID, message_id=val_msg_id,
                text=f"📌 طلب تحقق من الإصلاح:\nالمشكلة #{problem_id} - {problem['problem_text']}\nالمركبة: {problem['vehicle']}\nالسائق: ...\nالحالة: ✅ تم الإصلاح"
            )
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
    await query.edit_message_reply_markup(reply_markup=await run_db(build_problem_keyboard, problem_id))
    if updated_problem["media_type"] == "نظام" and new_ruglee == "تم الإصلاح":
        req_id = updated_problem.get("validation_requester") or updated_problem.get("user_id")
        if req_id:
            try:
                await context.bot.send_message(chat_id=req_id, text=f"تم تأكيد إصلاح الفيدانج للمركبة {updated_problem['vehicle']}. الرجاء إدخال الكيلومترات الحالية:")
                context.bot_data.setdefault("km_await", {})[req_id] = updated_problem["vehicle"]
            except Exception as e:
                logging.debug(f"Non-critical send/update failure: {e}")

async def fix_comment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("يجب إضافة تعليق أولاً.", show_alert=True)
    problem_id = int(query.data.split("_")[2])
    await context.bot.send_message(chat_id=query.from_user.id, text=f"يجب إضافة تعليق على المشكلة #{problem_id} أولاً.")
    context.user_data["awaiting_comment"] = problem_id

async def comment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    context.user_data["awaiting_comment"] = problem_id
    await context.bot.send_message(chat_id=query.from_user.id, text="📝 أرسل تعليقك على المشكلة:",
                                   reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input")]]))
    await query.answer("أرسل التعليق في المحادثة الخاصة.", show_alert=True)

async def delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("⛔ غير مصرح.", show_alert=True); return
    problem_id = int(query.data.split("_")[1])
    await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup([
        [InlineKeyboardButton("نعم", callback_data=f"confirmdel_{problem_id}"),
         InlineKeyboardButton("إلغاء", callback_data=f"cancel_{problem_id}")]
    ]))

async def confirm_delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    problem_id = int(query.data.split("_")[1])
    problem = await run_db(get_problem, problem_id)
    await run_db(delete_problem, problem_id)
    try:
        if problem and problem["media_type"] and problem["media_type"] != "نظام":
            await query.edit_message_caption(caption="🗑️ تم حذف المشكلة.")
        else:
            await query.edit_message_text("🗑️ تم حذف المشكلة.")
    except Exception as e:
        logging.warning(f"Problem #{problem_id} deleted from DB but could not edit its group message: {e}")

async def cancel_delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    problem = get_problem(problem_id)
    if not problem:
        await query.edit_message_reply_markup(reply_markup=None)
        return
    await query.edit_message_reply_markup(reply_markup=build_problem_keyboard(problem_id))

async def validation_request_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    problem = get_problem(problem_id)
    if not problem: return await query.answer("غير موجود.")
    if problem["validation_requester"] != 0:
        await query.answer("تم إرسال طلب تحقق سابقاً.", show_alert=True)
        return
    update_problem_status(problem_id, validation_requester=query.from_user.id)
    driver = get_driver(query.from_user.id)
    driver_name = driver["name"] if driver else "Unknown"
    msg_text = f"📌 طلب تحقق من الإصلاح:\nالمشكلة #{problem_id} - {problem['problem_text']}\nالمركبة: {problem['vehicle']}\nالسائق: {driver_name}\nالحالة: 📌 في انتظار التحقق"
    sent_msg = await context.bot.send_message(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_VALIDATION, text=msg_text,
                                              reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✅ تأكيد الإصلاح", callback_data=f"valrug_{problem_id}")]]))
    context.bot_data.setdefault("validation_msgs", {})[problem_id] = sent_msg.message_id
    await query.edit_message_text("تم إرسال طلب التحقق.")

async def valrug_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[1])
    problem = get_problem(problem_id)
    if not problem: return await query.answer("غير موجود.")
    if problem["media_type"] and not problem["comments"]:
        await query.answer("يجب إضافة تعليق أولاً.", show_alert=True)
        await context.bot.send_message(chat_id=query.from_user.id, text="يجب إضافة تعليق للمشكلة قبل وضعها كمُصلحة. أرسل التعليق هنا.")
        context.user_data["awaiting_comment"] = problem_id
        return
    new_ruglee = "تم الإصلاح" if problem["ruglee"] == "غير مُصلح" else "غير مُصلح"
    await run_db(update_problem_status, problem_id, ruglee=new_ruglee)
    updated_problem = await run_db(get_problem, problem_id)
    await _update_problem_message(updated_problem)
    post_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔒 إغلاق الطلب", callback_data=f"close_val_{problem_id}"),
         InlineKeyboardButton("🔄 إعادة فتح", callback_data=f"reopen_val_{problem_id}")]
    ])
    try:
        await query.edit_message_reply_markup(reply_markup=post_markup)
    except Exception as e:
        logging.warning(f"Failed to update val msg markup: {e}")
    context.bot_data.get("validation_msgs", {}).pop(problem_id, None)
    if updated_problem["media_type"] == "نظام" and new_ruglee == "تم الإصلاح":
        req_id = updated_problem.get("validation_requester") or updated_problem.get("user_id")
        if req_id:
            try:
                await context.bot.send_message(chat_id=req_id, text=f"تم تأكيد إصلاح الفيدانج للمركبة {updated_problem['vehicle']}. الرجاء إدخال الكيلومترات الحالية:")
                context.bot_data.setdefault("km_await", {})[req_id] = updated_problem["vehicle"]
            except Exception as e:
                logging.debug(f"Non-critical send/update failure: {e}")

async def close_val_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[2])
    problem = get_problem(problem_id)
    if not problem: return await query.edit_message_text("المشكلة غير موجودة.")
    try:
        await query.message.delete()
    except Exception as e:
        logging.warning(f"Could not delete val msg: {e}")
    driver_id = problem.get("validation_requester") or problem.get("user_id")
    if driver_id:
        try:
            await context.bot.send_message(chat_id=driver_id, text=f"تم إغلاق طلب التحقق للمشكلة #{problem_id}.")
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")

async def reopen_val_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    problem_id = int(query.data.split("_")[2])
    problem = get_problem(problem_id)
    if not problem: return await query.edit_message_text("المشكلة غير موجودة.")
    update_problem_status(problem_id, ruglee="غير مُصلح")
    driver_id = problem.get("validation_requester") or problem.get("user_id")
    if driver_id:
        try:
            await context.bot.send_message(chat_id=driver_id, text=f"تم إعادة فتح طلب التحقق للمشكلة #{problem_id}. الرجاء مراجعة الإصلاح.")
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
    orig_markup = InlineKeyboardMarkup([[InlineKeyboardButton("✅ تأكيد الإصلاح", callback_data=f"valrug_{problem_id}")]])
    try:
        await query.edit_message_reply_markup(reply_markup=orig_markup)
    except Exception as e:
        logging.debug(f"Non-critical send/update failure: {e}")

# Vidange confirm / modify
async def vidange_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("⛔ غير مصرح.", show_alert=True); return
    data = query.data.split("_")
    user_id = int(data[1])
    km = int(data[2])
    pending = context.bot_data.get("pending_vidange", {})
    info = pending.pop(query.message.message_id, None)
    if not info:
        await query.answer("انتهت صلاحية الطلب.", show_alert=True)
        return
    vehicle = info["vehicle"]
    add_km_reading(vehicle, km, driver_name=info.get("driver_name", ""))
    set_last_vidange_km(vehicle, km)
    await query.edit_message_text(f"✅ تم تأكيد الفيدانج للمركبة {vehicle} بقيمة {km} كم.")
    try:
        await context.bot.send_message(chat_id=user_id, text=f"✅ تم اعتماد تحديث الفيدانج للمركبة {vehicle}: {km} كم.")
    except Exception as e:
        logging.debug(f"Non-critical send/update failure: {e}")

async def vidange_modify_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("⛔ غير مصرح.", show_alert=True); return
    data = query.data.split("_")
    user_id = int(data[1])
    pending = context.bot_data.get("pending_vidange", {})
    info = pending.get(query.message.message_id)
    if not info:
        await query.answer("انتهت صلاحية الطلب.", show_alert=True)
        return
    context.user_data["vidange_modify"] = {
        "user_id": user_id,
        "vehicle": info["vehicle"],
        "message_id": query.message.message_id
    }
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
         InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])
    await query.edit_message_text("✏️ أرسل القيمة الصحيحة للكيلومتر بعد الفيدانج:", reply_markup=markup)

# Approval / Rejection callbacks
async def _is_admin_user(user_id: int) -> bool:
    """Accept both legacy ADMIN_IDS admins and admins persisted in PostgreSQL."""
    if user_id in ADMIN_IDS:
        return True
    try:
        return bool(await run_db(get_admin, user_id))
    except Exception:
        logging.exception("Could not verify admin %s", user_id)
        return False

async def approve_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not await _is_admin_user(query.from_user.id):
        await query.answer("⛔ غير مصرح لك.", show_alert=True)
        return
    await query.answer("جارٍ اعتماد السائق…")
    try:
        user_id = int(query.data.split("_", 1)[1])
    except (IndexError, ValueError):
        await query.answer("طلب اعتماد غير صالح.", show_alert=True)
        return

    await run_db(add_allowed_user, user_id, "approved")
    await run_db(set_driver, user_id, approval_status="approved", state="vehicle_selection")
    context.bot_data.setdefault("approval_msg", {}).pop(user_id, None)
    try:
        await query.edit_message_text(f"✅ تم قبول المستخدم {user_id}\nالحالة: تمت الموافقة، بانتظار اختيار المركبة.")
    except Exception as e:
        logging.warning("Could not update approval message for %s: %s", user_id, e)
    try:
        vehicles = await run_db(get_all_vehicles)
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ تم قبول تسجيلك. اختر المركبة المرتبطة بك:",
            reply_markup=vehicle_inline_keyboard(vehicles, "selv_"),
        )
    except Exception as e:
        logging.warning("Could not notify approved driver %s: %s", user_id, e)

async def reject_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not await _is_admin_user(query.from_user.id):
        await query.answer("⛔ غير مصرح لك.", show_alert=True)
        return
    await query.answer("جارٍ رفض الطلب…")
    try:
        user_id = int(query.data.split("_", 1)[1])
    except (IndexError, ValueError):
        await query.answer("طلب رفض غير صالح.", show_alert=True)
        return

    await run_db(set_driver, user_id, approval_status="rejected", state="idle")
    await run_db(add_allowed_user, user_id, "rejected")
    context.bot_data.setdefault("approval_msg", {}).pop(user_id, None)
    try:
        await query.edit_message_text(f"❌ تم رفض المستخدم {user_id}\nالحالة: مرفوض")
    except Exception as e:
        logging.warning("Could not update rejection message for %s: %s", user_id, e)
    try:
        await context.bot.send_message(chat_id=user_id, text="عذراً، تم رفض طلب التسجيل. يمكنك التواصل مع الإدارة.")
    except Exception as e:
        logging.warning("Could not notify rejected driver %s: %s", user_id, e)

# Admin submenu callbacks
async def admin_main_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("🕹️ لوحة التحكم:", reply_markup=admin_main_keyboard())

async def admin_vehicles_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("🚘 تعديل المركبات:", reply_markup=admin_vehicles_keyboard())

async def admin_drivers_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("👤 إدارة السائقين:", reply_markup=admin_drivers_keyboard())

async def admin_vidange_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("🛢️ إدارة الفيدانج:", reply_markup=admin_vidange_menu_keyboard())

async def admin_export_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("📋 تصدير Excel:", reply_markup=admin_export_menu_keyboard())

async def admin_approve_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT user_id, name FROM drivers WHERE approval_status='pending'")
        pending = cur.fetchall()
    if not pending:
        await query.edit_message_text("لا يوجد سائقون بانتظار القبول.")
        return
    rows = []
    for d in pending:
        label = d['name'] or str(d['user_id'])
        rows.append([
            InlineKeyboardButton(f"✅ قبول {label}", callback_data=f"approve_{d['user_id']}"),
            InlineKeyboardButton(f"❌ رفض {label}", callback_data=f"reject_{d['user_id']}"),
        ])
    await query.edit_message_text("اختر إجراءً للسائق المعلّق:", reply_markup=InlineKeyboardMarkup(rows))

async def admin_remove_driver_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    drivers = get_all_drivers()
    if not drivers:
        await query.edit_message_text("لا يوجد سائقون مسجلون.")
        return
    buttons = [InlineKeyboardButton(f"{d['name']} ({d['user_id']})", callback_data=f"rmdriver_{d['user_id']}") for d in drivers]
    await query.edit_message_text("اختر سائقًا لحذفه:", reply_markup=InlineKeyboardMarkup([buttons[i:i+2] for i in range(0, len(buttons), 2)]))

async def admin_drivers_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    drivers = get_all_drivers()
    if not drivers:
        await query.edit_message_text("لا يوجد سائقون نشطون.")
        return
    text = "👤 السائقون النشطون:\n" + "\n".join(f"• {d['name']} ({d['user_id']})" for d in drivers)
    await query.edit_message_text(text)

async def confirm_remove_driver(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = int(query.data.split("_")[1])
    await query.edit_message_text(f"هل أنت متأكد من حذف السائق {user_id}؟", reply_markup=InlineKeyboardMarkup([
        [InlineKeyboardButton("نعم", callback_data=f"confirmrm_{user_id}"),
         InlineKeyboardButton("إلغاء", callback_data="admin_drivers")]
    ]))

async def confirm_remove_driver_exec(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    user_id = int(query.data.split("_")[1])
    remove_driver(user_id)
    await query.edit_message_text(f"✅ تم حذف السائق {user_id}.")
    try:
        await context.bot.send_message(chat_id=user_id, text="تم إلغاء صلاحيتك لاستخدام البوت.")
    except Exception as e:
        logging.debug(f"Non-critical send/update failure: {e}")

# Admin input handler
async def admin_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS: return
    text = update.message.text.strip()
    if context.user_data.get("vidange_modify"):
        info = context.user_data.pop("vidange_modify")
        if not text.isdigit():
            await update.message.reply_text("يجب أن يكون الكيلومتر رقماً. حاول مرة أخرى:",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
                                                                               InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]]))
            context.user_data["vidange_modify"] = info
            return
        km = int(text)
        vehicle = info["vehicle"]
        last_km = get_latest_km(vehicle)
        if last_km is not None and km <= last_km:
            await update.message.reply_text(f"⚠️ الكيلومتر يجب أن يكون أكبر من آخر قراءة ({last_km} كم). أعد الإدخال.",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
                                                                               InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]]))
            context.user_data["vidange_modify"] = info
            return
        add_km_reading(vehicle, km, driver_name="مشرف")
        set_last_vidange_km(vehicle, km)
        try:
            await context.bot.edit_message_text(
                chat_id=ADMIN_GROUP_ID,
                message_id=info["message_id"],
                text=f"✅ تم تعديل الفيدانج للمركبة {vehicle} إلى {km} كم."
            )
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
        try:
            await context.bot.send_message(chat_id=info["user_id"], text=f"✅ تم تحديث الفيدانج للمركبة {vehicle} بقيمة {km} كم (بعد المراجعة).")
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
        await update.message.reply_text(f"✅ تم تحديث الفيدانج لـ {vehicle} = {km} كم.")
        return
    if context.user_data.get("admin_urgentvid"):
        context.user_data.pop("admin_urgentvid")
        parts = text.split()
        if len(parts) != 2 or not parts[1].isdigit():
            await update.message.reply_text("صيغة خاطئة. استخدم: CODE KM",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
                                                                               InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]]))
            return
        code = parts[0].upper()
        km = int(parts[1])
        if code not in get_all_vehicles():
            await update.message.reply_text("المركبة غير موجودة.")
            return
        set_last_vidange_km(code, km)
        latest = get_latest_km(code)
        if latest is not None and latest >= km + 9000:
            vidange_problem_id = add_problem(0, "نظام (عاجل)", code, f"Vidange عاجل {code}", "نظام")
            await context.bot.send_message(
                chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_VIDANGE,
                text=f"🚨 فيدانج عاجل: المركبة {code}\nالعداد الحالي: {latest} كم\nآخر فيدانج (محدث): {km} كم\n⚪ الحالة: قيد الانتظار",
                reply_markup=build_problem_keyboard(vidange_problem_id)
            )
        await update.message.reply_text(f"✅ تم تعيين آخر فيدانج عاجل للمركبة {code} = {km} كم.")
        return
    if context.user_data.get("admin_add_veh"):
        context.user_data.pop("admin_add_veh")
        add_vehicle(text.upper())
        await update.message.reply_text(f"✅ تمت إضافة {text.upper()}.")
    elif context.user_data.get("admin_rem_veh"):
        context.user_data.pop("admin_rem_veh")
        remove_vehicle(text.upper())
        await update.message.reply_text(f"🗑️ تم حذف {text.upper()}.")
    elif context.user_data.get("admin_setvid"):
        context.user_data.pop("admin_setvid")
        parts = text.split()
        if len(parts) != 2 or not parts[1].isdigit():
            await update.message.reply_text("صيغة خاطئة. استخدم: CODE KM",
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
                                                                               InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]]))
            return
        code = parts[0].upper()
        km = int(parts[1])
        if code not in get_all_vehicles():
            await update.message.reply_text("المركبة غير موجودة.")
            return
        set_last_vidange_km(code, km)
        await update.message.reply_text(f"✅ تم تعيين آخر فيدانج لـ {code} = {km} كم.")

# Settings callbacks
async def settings_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    videos = get_all_help_videos()
    if not videos:
        await query.edit_message_text("لا يوجد فيديو تعليمي حالياً.")
        return
    for v in videos:
        try:
            await context.bot.send_video(chat_id=query.from_user.id, video=v["file_id"], caption=v.get("description"))
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
    await query.edit_message_text("تم إرسال الفيديوهات التعليمية.")

async def settings_history_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    problems = await run_db(get_driver_problems, user_id)
    if not problems:
        await query.edit_message_text("لا توجد شكاوي مسجلة.")
        return
    text = "📜 سجل شكاويي:\n"
    for p in problems[:10]:
        text += f"#{p['id']} | {p['date']} | {p['problem_text'][:30]} | {status_icon_and_text(p)}\n"
    await query.edit_message_text(safe_text(text))

async def settings_change_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    set_driver(query.from_user.id, state="name_entry")
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("إلغاء", callback_data="cancel_name")]])
    await query.edit_message_text("أرسل اسمك الجديد:", reply_markup=markup)

async def cancel_name_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    set_driver(query.from_user.id, state="idle")
    await query.edit_message_text("تم إلغاء تغيير الاسم.")

# NEW: Change vehicle settings callback
async def settings_change_vehicle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    vehicles = get_all_vehicles()
    await query.edit_message_text("اختر مركبتك الجديدة:", reply_markup=vehicle_inline_keyboard(vehicles, "selv_"))

# Help video commands
async def set_help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS: return
    if not update.message.video and not update.message.reply_to_message:
        await update.message.reply_text("أرسل الفيديو (كملف فيديو) أو رد على فيديو.")
        return
    if update.message.reply_to_message and update.message.reply_to_message.video:
        file_id = update.message.reply_to_message.video.file_id
        desc = update.message.text or ""
        add_help_video(file_id, desc)
        await update.message.reply_text("تم حفظ الفيديو بنجاح.")
    elif update.message.video:
        file_id = update.message.video.file_id
        desc = update.message.caption or ""
        add_help_video(file_id, desc)
        await update.message.reply_text("تم حفظ الفيديو بنجاح.")
    else:
        await update.message.reply_text("الرجاء إرسال فيديو أو الرد على فيديو.")

async def remove_help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS: return
    videos = get_all_help_videos()
    if not videos:
        await update.message.reply_text("لا توجد فيديوهات للحذف.")
        return
    buttons = []
    for v in videos:
        desc = v["description"] or f"فيديو {v['id']}"
        buttons.append([InlineKeyboardButton(f"🗑️ {desc}", callback_data=f"delhelp_{v['id']}")])
    await update.message.reply_text("اختر فيديو للحذف:", reply_markup=InlineKeyboardMarkup(buttons))

async def delete_help_video_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    video_id = int(query.data.split("_")[1])
    delete_help_video(video_id)
    await query.edit_message_text("✅ تم حذف الفيديو.")

# Broadcast command (super admin only)
async def broadcast_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS: return
    if not context.args:
        await update.message.reply_text("استخدم: /broadcast <النص>")
        return
    message = " ".join(context.args)
    drivers = await run_db(get_all_drivers)
    count = 0
    for d in drivers:
        try:
            await context.bot.send_message(chat_id=d["user_id"], text=message)
            count += 1
        except Exception as e:
            logging.debug(f"Non-critical send/update failure: {e}")
    await update.message.reply_text(f"تم إرسال الرسالة إلى {count} سائق.")

async def delete_problem_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/delete <رمز المركبة> <رقم المشكلة> — يحذف الشكوى نهائياً من قاعدة البيانات
    (وبالتالي لن تظهر في أي تصدير Excel لاحق، لأن الملفات تُبنى دائماً من قاعدة
    البيانات الحية وقت الطلب ولا يوجد جدول Excel مخزَّن بشكل دائم)."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ هذا الأمر مخصص للمشرفين فقط.")
        return
    if len(context.args) != 2:
        await update.message.reply_text("استخدم: /delete <رمز المركبة> <رقم المشكلة>\nمثال: /delete M15 42")
        return
    vehicle_code = context.args[0].upper()
    try:
        problem_id = int(context.args[1])
    except ValueError:
        await update.message.reply_text("رقم المشكلة يجب أن يكون رقماً صحيحاً.")
        return

    problem = await run_db(get_problem, problem_id)
    if not problem:
        await update.message.reply_text(f"لا توجد شكوى برقم #{problem_id}.")
        return
    if problem["vehicle"] != vehicle_code:
        await update.message.reply_text(
            f"⚠️ الشكوى #{problem_id} تخص المركبة {problem['vehicle']} وليس {vehicle_code}. "
            "تحقق من الرمز والرقم ثم أعد المحاولة (لمنع حذف شكوى بالخطأ)."
        )
        return

    await run_db(delete_problem, problem_id)

    # محاولة تحديث/حذف الرسالة الأصلية في المجموعة الإدارية إن وُجدت
    if problem.get("group_message_id"):
        try:
            if problem["media_type"] and problem["media_type"] != "نظام":
                await context.bot.edit_message_caption(
                    chat_id=ADMIN_GROUP_ID, message_id=problem["group_message_id"],
                    caption="🗑️ تم حذف هذه المشكلة عبر أمر /delete."
                )
            else:
                await context.bot.edit_message_text(
                    chat_id=ADMIN_GROUP_ID, message_id=problem["group_message_id"],
                    text="🗑️ تم حذف هذه المشكلة عبر أمر /delete."
                )
        except Exception as e:
            logging.warning(f"Problem #{problem_id} deleted via /delete but could not edit its group message: {e}")

    await update.message.reply_text(f"✅ تم حذف الشكوى #{problem_id} الخاصة بالمركبة {vehicle_code} نهائياً من قاعدة البيانات.")

async def applogin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/applogin — يولّد كوداً من 6 أرقام صالحاً لمدة 10 دقائق لتسجيل الدخول
    في تطبيق أندرويد الإداري. يعمل فقط للمستخدمين المسجّلين في جدول admins."""
    user_id = update.effective_user.id
    admin = await run_db(get_admin, user_id)
    if not admin:
        await update.message.reply_text("⛔ حسابك غير مسجَّل كمشرف في النظام.")
        return
    code = await run_db(create_login_code, user_id)
    role_ar = "مشرف عام (Super Admin)" if admin["role"] == "super_admin" else "مشرف (Admin)"
    await update.message.reply_text(
        f"🔑 كود تسجيل الدخول لتطبيق أندرويد:\n\n<code>{code}</code>\n\n"
        f"الصلاحية: {role_ar}\n"
        f"⏳ صالح لمدة 10 دقائق فقط.\n"
        f"أدخل معرّفك (ID: <code>{user_id}</code>) وهذا الكود في شاشة تسجيل الدخول بالتطبيق.",
        parse_mode="HTML"
    )

async def fixkm_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/fixkm <رمز المركبة> <القيمة الصحيحة> — يصحّح آخر قراءة عداد أُدخلت
    بالخطأ (بدل حذفها، حفاظاً على تاريخ القراءات). يُستخدم مثلاً عندما يُدخل
    السائق 165000 بدل 16500 بالخطأ."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ هذا الأمر مخصص للمشرفين فقط.")
        return
    if len(context.args) != 2:
        await update.message.reply_text(
            "استخدم: /fixkm <رمز المركبة> <القيمة الصحيحة>\n"
            "مثال: /fixkm M15 16500\n\n"
            "يُستخدم لتصحيح آخر قراءة عداد أُدخلت بالخطأ (مثل 165000 بدل 16500)."
        )
        return
    vehicle_code = context.args[0].upper()
    try:
        corrected_km = int(context.args[1])
    except ValueError:
        await update.message.reply_text("القيمة الصحيحة يجب أن تكون رقماً صحيحاً.")
        return
    if corrected_km <= 0:
        await update.message.reply_text("القيمة يجب أن تكون أكبر من صفر.")
        return

    all_vehicles = await run_db(get_all_vehicles)
    if vehicle_code not in all_vehicles:
        await update.message.reply_text(f"لا توجد مركبة بالرمز {vehicle_code}.")
        return

    success, old_km = await run_db(fix_latest_km_reading, vehicle_code, corrected_km)
    if not success:
        await update.message.reply_text(f"لا توجد أي قراءة عداد مسجَّلة أصلاً للمركبة {vehicle_code} لتصحيحها.")
        return

    await update.message.reply_text(
        f"✅ تم تصحيح آخر قراءة عداد للمركبة {vehicle_code}:\n"
        f"القيمة الخاطئة السابقة: {old_km} كم\n"
        f"القيمة الصحيحة الجديدة: {corrected_km} كم"
    )

# Dashboard functions
async def _send_dashboard(chat_id: int, thread_id: int = None, bot=None):
    """Build and send the dashboard without blocking Telegram's event loop."""
    sender = bot or app.bot
    try:
        vehicles = await run_db(get_all_vehicles)
        if not vehicles:
            logging.warning("Dashboard skipped: no vehicles found")
            return False
        buttons = []
        for vehicle in vehicles:
            text = await run_db(dashboard_button_text, vehicle)
            buttons.append(InlineKeyboardButton(text, callback_data=f"hist_{vehicle}"))
        markup = InlineKeyboardMarkup([buttons[i:i+2] for i in range(0, len(buttons), 2)])
        send_kwargs = {
            "chat_id": chat_id,
            "text": "📊 الحالة اليومية للمركبات:",
            "reply_markup": markup,
        }
        if thread_id:
            send_kwargs["message_thread_id"] = thread_id
        await sender.send_message(**send_kwargs)
        return True
    except Exception:
        logging.exception("Dashboard send failed for chat_id=%s thread_id=%s", chat_id, thread_id)
        return False

async def dashboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    thread_id = update.message.message_thread_id if update.message else None
    await _send_dashboard(chat_id, thread_id)
    if update.message:
        await update.message.reply_text("تم إرسال لوحة القيادة.")

# FIXED: admin_dash now uses the thread_id from the callback message, falling back to TOPIC_GENERAL
async def admin_dash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    thread_id = query.message.message_thread_id if query.message else None
    if not thread_id:
        thread_id = TOPIC_GENERAL
    await _send_dashboard(ADMIN_GROUP_ID, thread_id)
    await query.edit_message_text("تم إرسال لوحة القيادة.")

async def scheduled_dashboard(context: ContextTypes.DEFAULT_TYPE):
    logging.info("Running scheduled daily dashboard")
    await _send_dashboard(ADMIN_GROUP_ID, TOPIC_GENERAL, bot=context.bot)

async def weekly_excel(context: ContextTypes.DEFAULT_TYPE):
    file = await run_db(generate_problems_excel)
    try:
        await context.bot.send_document(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_GENERAL, document=file, filename="المشاكل_الأسبوعي.xlsx")
    except Exception as e:
        logging.warning(f"Weekly excel send failed: {e}")

def _next_run_with_grace(now: datetime, target_dt: datetime, interval: timedelta, grace_minutes: int = 10) -> datetime:
    """يحسب موعد أول تشغيل لمهمة مجدولة، مع فترة سماح للتعويض عند إعادة التشغيل:
    - إن كان الموعد لم يحن بعد اليوم -> يُجدوَل بشكل طبيعي في موعده.
    - إن تجاوز الموعد بأقل من grace_minutes دقيقة (مثلاً بسبب إعادة تشغيل البوت
      قريباً من 7:30) -> تُرسَل المهمة فوراً تعويضاً عن الفترة الفائتة.
    - إن تجاوز الموعد بأكثر من grace_minutes دقيقة -> تُؤجَّل للدورة القادمة
      (غداً للمهمة اليومية، الأسبوع القادم للمهمة الأسبوعية) بدل إرسالها متأخرة جداً.
    """
    grace = timedelta(minutes=grace_minutes)
    if now < target_dt:
        return target_dt
    elif now <= target_dt + grace:
        return now + timedelta(seconds=5)
    else:
        return target_dt + interval

def schedule_jobs(app: Application):
    """Register jobs after Application.initialize(), when JobQueue is ready."""
    if app.job_queue is None:
        raise RuntimeError(
            "JobQueue is unavailable. Install python-telegram-bot[job-queue] in requirements.txt."
        )
    now = datetime.now(TZ)
    target = time(7, 30, 0)

    today_target = datetime.combine(now.date(), target, tzinfo=TZ)
    next_daily = _next_run_with_grace(now, today_target, timedelta(days=1))
    daily_job = app.job_queue.run_repeating(
        scheduled_dashboard,
        interval=timedelta(days=1),
        first=next_daily,
        name="daily_dashboard",
    )

    days_until_sat = (5 - now.weekday()) % 7
    today_or_next_sat_target = datetime.combine(
        now.date() + timedelta(days=days_until_sat), target, tzinfo=TZ
    )
    next_sat = _next_run_with_grace(now, today_or_next_sat_target, timedelta(days=7))
    weekly_job = app.job_queue.run_repeating(
        weekly_excel,
        interval=timedelta(days=7),
        first=next_sat,
        name="weekly_excel",
    )
    logging.info(
        "Scheduled jobs registered: daily_dashboard first=%s, weekly_excel first=%s",
        next_daily.isoformat(),
        next_sat.isoformat(),
    )
    return daily_job, weekly_job

# Export functions
async def export_problems(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ هذا الأمر مخصص للمشرفين فقط.")
        return
    file = await run_db(generate_problems_excel)
    if update.message:
        await update.message.reply_document(document=file, filename="المشاكل.xlsx")
    else:
        await context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename="المشاكل.xlsx")

async def export_vidange(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ هذا الأمر مخصص للمشرفين فقط.")
        return
    file = await run_db(generate_vidange_excel)
    if update.message:
        await update.message.reply_document(document=file, filename="الفيدانج.xlsx")
    else:
        await context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename="الفيدانج.xlsx")

async def export_vidange_vehicle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ هذا الأمر مخصص للمشرفين فقط.")
        return
    code = context.args[0].upper() if context.args else None
    all_vehicles = await run_db(get_all_vehicles)
    if not code or code not in all_vehicles:
        await update.message.reply_text("استخدم: /vidange <CODE> مع رمز مركبة صحيح.")
        return
    file = await run_db(generate_vidange_excel, code)
    await update.message.reply_document(document=file, filename=f"فيدانج_{code}.xlsx")

# Sticky admin panel command
async def set_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    await update.message.reply_text(
        "تم تفعيل لوحة التحكم السريعة. استخدم الزر أدناه للوصول السريع.",
        reply_markup=ADMIN_STICKY_KEYBOARD
    )

async def remove_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    await update.message.reply_text(
        "تم إخفاء لوحة التحكم السريعة.",
        reply_markup=ReplyKeyboardRemove()
    )

# Admin action callbacks (addveh, remveh, etc.)
async def admin_addveh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
         InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])
    await query.edit_message_text("أرسل رمز المركبة الجديدة:", reply_markup=markup)
    context.user_data["admin_add_veh"] = True

async def admin_remveh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
         InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])
    await query.edit_message_text("أرسل رمز المركبة المراد حذفها:", reply_markup=markup)
    context.user_data["admin_rem_veh"] = True

async def admin_setvid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
         InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])
    await query.edit_message_text("أرسل رمز المركبة ثم الكيلومتر (مثال: M02 150000):", reply_markup=markup)
    context.user_data["admin_setvid"] = True

async def admin_urgentvid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("إلغاء", callback_data="cancel_input"),
         InlineKeyboardButton("↩️ رجوع", callback_data="admin_main")]
    ])
    await query.edit_message_text("أرسل رمز المركبة ثم الكيلومتر الجديد للفيدانج العاجل (مثال: M02 158000):", reply_markup=markup)
    context.user_data["admin_urgentvid"] = True

async def admin_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    file = await run_db(generate_problems_excel)
    await context.bot.send_document(chat_id=query.message.chat_id, document=file, filename="المشاكل.xlsx")
    await query.edit_message_text("تم إرسال ملف المشاكل.")

async def admin_vid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    file = await run_db(generate_vidange_excel)
    await context.bot.send_document(chat_id=query.message.chat_id, document=file, filename="الفيدانج.xlsx")
    await query.edit_message_text("تم إرسال ملف الفيدانج.")

async def admin_listveh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.from_user.id not in ADMIN_IDS: return
    vehicles = await run_db(get_all_vehicles)
    text = "🚘 المركبات المتاحة:\n" + "\n".join(f"• {v}" for v in vehicles) if vehicles else "لا توجد مركبات."
    await query.edit_message_text(safe_text(text))

async def vehicle_history_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    vehicle = query.data.split("_", 1)[1]
    history = await run_db(get_vehicle_history, vehicle)
    problems, readings = history["problems"], history["readings"]
    group_id_str = str(ADMIN_GROUP_ID)
    group_id_for_link = group_id_str[4:] if group_id_str.startswith("-100") else None
    text = f"🚘 تاريخ المركبة {vehicle}:\n"
    if problems:
        text += "\n📋 المشاكل:\n"
        for p in problems:
            if p['group_message_id'] and group_id_for_link:
                link = f"https://t.me/c/{group_id_for_link}/{p['group_message_id']}"
                label = f"<a href='{link}'>#{p['id']}</a>"
            else:
                label = f"#{p['id']}"
            text += f"  {label} | {p['date']} | {p['problem_text'][:40]} | {status_icon_and_text(p)}\n"
    else:
        text += "لا توجد مشاكل مسجلة.\n"
    if readings:
        text += "\n🛢️ آخر قراءات العداد:\n"
        for d, k in readings:
            text += f"  {d} - {k} كم\n"
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("تم", callback_data="done_hist")]])
    await context.bot.send_message(chat_id=ADMIN_GROUP_ID, message_thread_id=TOPIC_HISTORY, text=safe_text(text), parse_mode="HTML", reply_markup=markup)

async def done_history_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    try:
        await query.message.delete()
    except Exception as e:
        logging.warning(f"Could not delete history message: {e}")

# Command handlers for /admin, /panel, etc.
async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🕹️ لوحة التحكم:", reply_markup=admin_main_keyboard())

async def panel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    thread_id = update.message.message_thread_id if update.message else None
    keyboard = get_topic_keyboard(thread_id) if thread_id else None
    if keyboard:
        await update.message.reply_text("🕹️ لوحة التحكم الخاصة بهذا القسم:", reply_markup=keyboard)
    else:
        await update.message.reply_text("🕹️ لوحة التحكم الكاملة:", reply_markup=admin_main_keyboard())

# Error handler
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logging.error(msg="Exception while handling an update:", exc_info=context.error)

# Webhook and aiohttp
async def health(request):
    return web.Response(text="Bot is running")

async def telegram_webhook(request):
    if request.headers.get("X-Telegram-Bot-Api-Secret-Token") != WEBHOOK_SECRET:
        logging.warning("Rejected webhook request with invalid/missing secret token.")
        return web.Response(status=403)
    try:
        data = await request.json()
    except Exception as e:
        logging.warning(f"Invalid webhook JSON payload: {e}")
        return web.Response(status=400)
    try:
        update = Update.de_json(data, app.bot)
        await app.process_update(update)
    except Exception as e:
        logging.error(f"Error processing update: {e}", exc_info=True)
    return web.Response(status=200)

async def set_webhook(app: Application):
    webhook_url = f"{WEBHOOK_URL}/telegram"
    await app.bot.set_webhook(url=webhook_url, secret_token=WEBHOOK_SECRET)
    logging.info(f"Webhook set to {webhook_url}")

# ========================================================================
# REST API لتطبيق أندرويد الإداري (/api/*)
# ========================================================================

def issue_jwt(user_id: int, role: str) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "exp": datetime.now(TZ) + timedelta(days=JWT_EXPIRY_DAYS),
        "iat": datetime.now(TZ),
    }
    return pyjwt.encode(payload, APP_JWT_SECRET, algorithm="HS256")

def json_ok(data=None, **extra):
    body = {"ok": True}
    if data is not None:
        body["data"] = data
    body.update(extra)
    return web.json_response(body)

def json_err(message: str, status: int = 400):
    return web.json_response({"ok": False, "error": message}, status=status)

async def require_auth(request, min_role: str = "admin"):
    """يتحقق من رمز JWT في ترويسة Authorization: Bearer <token>.
    يرفع web.HTTPException مباشرة عند الفشل (401/403).
    min_role='super_admin' يمنع دخول المشرفين العاديين."""
    header = request.headers.get("Authorization", "")
    if not header.startswith("Bearer "):
        raise web.HTTPUnauthorized(text='{"ok": false, "error": "missing bearer token"}', content_type="application/json")
    token = header[7:]
    try:
        payload = pyjwt.decode(token, APP_JWT_SECRET, algorithms=["HS256"])
    except pyjwt.ExpiredSignatureError:
        raise web.HTTPUnauthorized(text='{"ok": false, "error": "token expired"}', content_type="application/json")
    except pyjwt.InvalidTokenError:
        raise web.HTTPUnauthorized(text='{"ok": false, "error": "invalid token"}', content_type="application/json")
    # التحقق من أن الحساب ما زال مشرفاً فعلياً (وليس محذوفاً منذ إصدار الرمز)
    admin = await run_db(get_admin, payload["user_id"])
    if not admin:
        raise web.HTTPUnauthorized(text='{"ok": false, "error": "account no longer an admin"}', content_type="application/json")
    if min_role == "super_admin" and admin["role"] != "super_admin":
        raise web.HTTPForbidden(text='{"ok": false, "error": "super_admin role required"}', content_type="application/json")
    return admin

# ---- المصادقة ----
async def api_auth_verify(request):
    try:
        body = await request.json()
    except Exception:
        return json_err("invalid JSON body")
    user_id = body.get("telegram_user_id")
    code = body.get("code")
    if not user_id or not code:
        return json_err("telegram_user_id و code مطلوبان")
    try:
        user_id = int(user_id)
    except (TypeError, ValueError):
        return json_err("telegram_user_id غير صالح")
    admin = await run_db(get_admin, user_id)
    if not admin:
        return json_err("هذا الحساب غير مسجَّل كمشرف", status=403)
    valid = await run_db(verify_and_consume_login_code, user_id, str(code))
    if not valid:
        return json_err("الكود غير صحيح أو منتهي الصلاحية", status=401)
    token = issue_jwt(user_id, admin["role"])
    return json_ok({"token": token, "role": admin["role"], "display_name": admin["display_name"], "user_id": user_id})

# ---- لوحة القيادة ----
async def api_dashboard(request):
    await require_auth(request)
    vehicles = await run_db(get_all_vehicles)
    result = []
    for v in vehicles:
        info = await run_db(get_vehicle_cache_entry, v)
        result.append({
            "vehicle": v,
            "status": info["status"],  # bad | en_cours | vidange | good
            "open_count": info["open_count"],
            "remaining_km": info["remaining_km"],
        })
    return json_ok(result)

# ---- المشاكل / الشكاوى ----
async def api_problems_list(request):
    await require_auth(request)
    vehicle = request.query.get("vehicle")
    status = request.query.get("status")
    problems = await run_db(get_all_problems)
    if vehicle:
        problems = [p for p in problems if p["vehicle"] == vehicle.upper()]
    if status:
        problems = [p for p in problems if p["status"] == status]
    return json_ok(problems)

async def api_problem_fix(request):
    admin = await require_auth(request)
    problem_id = int(request.match_info["id"])
    problem = await run_db(get_problem, problem_id)
    if not problem:
        return json_err("المشكلة غير موجودة", status=404)
    if problem["media_type"] and problem["media_type"] != "نظام" and not problem["comments"]:
        return json_err("يجب إضافة تعليق أولاً قبل تأكيد الإصلاح", status=409)
    new_ruglee = "تم الإصلاح" if problem["ruglee"] == "غير مُصلح" else "غير مُصلح"
    await run_db(update_problem_status, problem_id, None, new_ruglee)
    updated_problem = await run_db(get_problem, problem_id)
    await _update_problem_message(updated_problem)
    if updated_problem["media_type"] == "نظام" and new_ruglee == "تم الإصلاح":
        req_id = updated_problem.get("validation_requester") or updated_problem.get("user_id")
        if req_id:
            try:
                await app.bot.send_message(chat_id=req_id, text=f"تم تأكيد إصلاح الفيدانج للمركبة {updated_problem['vehicle']}. الرجاء إدخال الكيلومترات الحالية:")
                app.bot_data.setdefault("km_await", {})[req_id] = updated_problem["vehicle"]
            except Exception as e:
                logging.debug(f"Non-critical send failure: {e}")
    return json_ok({"problem_id": problem_id, "ruglee": new_ruglee}, by=admin["user_id"])

async def api_problem_comment(request):
    await require_auth(request)
    problem_id = int(request.match_info["id"])
    try:
        body = await request.json()
    except Exception:
        return json_err("invalid JSON body")
    text = (body.get("text") or "").strip()
    if not text:
        return json_err("النص مطلوب")
    problem = await run_db(get_problem, problem_id)
    if not problem:
        return json_err("المشكلة غير موجودة", status=404)
    await run_db(set_problem_comment, problem_id, text)
    return json_ok({"problem_id": problem_id})

async def api_problem_delete(request):
    admin = await require_auth(request)
    problem_id = int(request.match_info["id"])
    problem = await run_db(get_problem, problem_id)
    if not problem:
        return json_err("المشكلة غير موجودة", status=404)
    await run_db(delete_problem, problem_id)
    if problem.get("group_message_id"):
        try:
            if problem["media_type"] and problem["media_type"] != "نظام":
                await app.bot.edit_message_caption(chat_id=ADMIN_GROUP_ID, message_id=problem["group_message_id"], caption="🗑️ تم حذف هذه المشكلة عبر تطبيق أندرويد.")
            else:
                await app.bot.edit_message_text(chat_id=ADMIN_GROUP_ID, message_id=problem["group_message_id"], text="🗑️ تم حذف هذه المشكلة عبر تطبيق أندرويد.")
        except Exception as e:
            logging.warning(f"Problem #{problem_id} deleted via app but could not edit group message: {e}")
    return json_ok({"problem_id": problem_id}, deleted_by=admin["user_id"])

# ---- السائقون ----
async def api_drivers_list(request):
    await require_auth(request)
    status_filter = request.query.get("filter", "all")  # all | pending | approved
    drivers = await run_db(_fetch_drivers_by_filter, status_filter)
    return json_ok(drivers)

def _fetch_drivers_by_filter(status_filter: str) -> list:
    with db_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if status_filter == "pending":
            cur.execute("SELECT user_id, name, vehicle, approval_status FROM drivers WHERE approval_status='pending'")
        elif status_filter == "approved":
            cur.execute("SELECT user_id, name, vehicle, approval_status FROM drivers WHERE approval_status='approved'")
        else:
            cur.execute("SELECT user_id, name, vehicle, approval_status FROM drivers")
        return [dict(r) for r in cur.fetchall()]

async def api_driver_approve(request):
    await require_auth(request)
    user_id = int(request.match_info["user_id"])
    await run_db(add_allowed_user, user_id)
    await run_db(set_driver, user_id, approval_status="approved", state="vehicle_selection")
    try:
        await app.bot.send_message(chat_id=user_id, text="تم قبولك. يمكنك الآن اختيار مركبتك:")
        vehicles = await run_db(get_all_vehicles)
        await app.bot.send_message(chat_id=user_id, text="اختر مركبتك:", reply_markup=vehicle_inline_keyboard(vehicles, "selv_"))
    except Exception as e:
        logging.debug(f"Non-critical send failure: {e}")
    return json_ok({"user_id": user_id, "status": "approved"})

async def api_driver_reject(request):
    await require_auth(request)
    user_id = int(request.match_info["user_id"])
    await run_db(set_driver, user_id, approval_status="rejected")
    await run_db(add_allowed_user, user_id, status="rejected")
    try:
        await app.bot.send_message(chat_id=user_id, text="⛔ تم رفض طلبك من قبل الإدارة.")
    except Exception as e:
        logging.debug(f"Non-critical send failure: {e}")
    return json_ok({"user_id": user_id, "status": "rejected"})

async def api_driver_delete(request):
    await require_auth(request, min_role="super_admin")
    user_id = int(request.match_info["user_id"])
    await run_db(remove_driver, user_id)
    return json_ok({"user_id": user_id, "status": "removed"})

# ---- المركبات ----
async def api_vehicles_list(request):
    await require_auth(request)
    vehicles = await run_db(get_all_vehicles)
    return json_ok(vehicles)

async def api_vehicle_add(request):
    await require_auth(request, min_role="super_admin")
    try:
        body = await request.json()
    except Exception:
        return json_err("invalid JSON body")
    code = (body.get("code") or "").strip().upper()
    if not code or not code.isalnum() or len(code) > 10:
        return json_err("رمز المركبة يجب أن يكون أحرفاً/أرقاماً فقط (حتى 10 محارف)")
    await run_db(add_vehicle, code)
    return json_ok({"code": code})

async def api_vehicle_remove(request):
    await require_auth(request, min_role="super_admin")
    code = request.match_info["code"].upper()
    await run_db(remove_vehicle, code)
    return json_ok({"code": code})

# ---- الفيدانج ----
async def api_vidange_list(request):
    await require_auth(request)
    vehicles = await run_db(get_all_vehicles)
    result = []
    for v in vehicles:
        last_vid = await run_db(get_last_vidange_km, v)
        latest_km = await run_db(get_latest_km, v)
        result.append({
            "vehicle": v,
            "last_vidange_km": last_vid,
            "latest_km": latest_km,
            "remaining_km": (last_vid + 10000 - latest_km) if (latest_km is not None and last_vid) else None,
        })
    return json_ok(result)

async def api_vidange_set(request):
    await require_auth(request, min_role="super_admin")
    vehicle = request.match_info["vehicle"].upper()
    try:
        body = await request.json()
        km = int(body.get("km"))
    except Exception:
        return json_err("km مطلوب كرقم صحيح")
    await run_db(set_last_vidange_km, vehicle, km)
    return json_ok({"vehicle": vehicle, "last_vidange_km": km})

# ---- التصدير ----
async def api_export_problems(request):
    await require_auth(request)
    file = await run_db(generate_problems_excel)
    return web.Response(
        body=file.read(),
        headers={"Content-Disposition": "attachment; filename=\"المشاكل.xlsx\""},
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

async def api_export_vidange(request):
    await require_auth(request)
    vehicle = request.query.get("vehicle")
    file = await run_db(generate_vidange_excel, vehicle)
    return web.Response(
        body=file.read(),
        headers={"Content-Disposition": "attachment; filename=\"الفيدانج.xlsx\""},
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---- إدارة المشرفين (لمشرف عام فقط) ----
async def api_admins_list(request):
    await require_auth(request, min_role="super_admin")
    return json_ok(await run_db(list_admins))

async def api_admin_add(request):
    await require_auth(request, min_role="super_admin")
    try:
        body = await request.json()
    except Exception:
        return json_err("invalid JSON body")
    user_id = body.get("telegram_user_id")
    role = body.get("role", "admin")
    display_name = body.get("display_name", "")
    if role not in ("admin", "super_admin"):
        return json_err("role يجب أن تكون admin أو super_admin")
    try:
        user_id = int(user_id)
    except (TypeError, ValueError):
        return json_err("telegram_user_id غير صالح")
    await run_db(upsert_admin, user_id, role, display_name)
    return json_ok({"user_id": user_id, "role": role})

async def api_admin_remove(request):
    admin = await require_auth(request, min_role="super_admin")
    user_id = int(request.match_info["user_id"])
    if user_id == admin["user_id"]:
        return json_err("لا يمكنك حذف حسابك الخاص")
    await run_db(remove_admin, user_id)
    return json_ok({"user_id": user_id, "status": "removed"})

@web.middleware
async def api_error_middleware(request, handler):
    try:
        return await handler(request)
    except web.HTTPException:
        raise
    except Exception as e:
        logging.error(f"Unhandled /api error on {request.path}: {e}", exc_info=True)
        return json_err("internal server error", status=500)

def register_api_routes(api: web.Application):
    api.router.add_post("/api/auth/verify", api_auth_verify)
    api.router.add_get("/api/dashboard", api_dashboard)
    api.router.add_get("/api/problems", api_problems_list)
    api.router.add_post("/api/problems/{id}/fix", api_problem_fix)
    api.router.add_post("/api/problems/{id}/comment", api_problem_comment)
    api.router.add_delete("/api/problems/{id}", api_problem_delete)
    api.router.add_get("/api/drivers", api_drivers_list)
    api.router.add_post("/api/drivers/{user_id}/approve", api_driver_approve)
    api.router.add_post("/api/drivers/{user_id}/reject", api_driver_reject)
    api.router.add_delete("/api/drivers/{user_id}", api_driver_delete)
    api.router.add_get("/api/vehicles", api_vehicles_list)
    api.router.add_post("/api/vehicles", api_vehicle_add)
    api.router.add_delete("/api/vehicles/{code}", api_vehicle_remove)
    api.router.add_get("/api/vidange", api_vidange_list)
    api.router.add_post("/api/vidange/{vehicle}/set", api_vidange_set)
    api.router.add_get("/api/export/problems", api_export_problems)
    api.router.add_get("/api/export/vidange", api_export_vidange)
    api.router.add_get("/api/admins", api_admins_list)
    api.router.add_post("/api/admins", api_admin_add)
    api.router.add_delete("/api/admins/{user_id}", api_admin_remove)

async def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    init_db()
    global app
    persistence = PicklePersistence(filepath=PERSISTENCE_PATH)
    app = Application.builder().token(BOT_TOKEN).persistence(persistence).build()

    # Command handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("admin", admin_command))
    app.add_handler(CommandHandler("panel", panel_command))
    app.add_handler(CommandHandler("dashboard", dashboard_command))
    app.add_handler(CommandHandler("sethelp", set_help_cmd))
    app.add_handler(CommandHandler("removehelp", remove_help_cmd))
    app.add_handler(CommandHandler("broadcast", broadcast_cmd))
    app.add_handler(CommandHandler("delete", delete_problem_cmd))
    app.add_handler(CommandHandler("applogin", applogin_cmd))
    app.add_handler(CommandHandler("fixkm", fixkm_cmd))
    app.add_handler(CommandHandler("export", export_problems))
    app.add_handler(CommandHandler("export_vidange", export_vidange))
    app.add_handler(CommandHandler("vidange", export_vidange_vehicle))
    app.add_handler(CommandHandler("setadminpanel", set_admin_panel))
    app.add_handler(CommandHandler("removeadminpanel", remove_admin_panel))

    # Text / Media handlers (restricted to private chats)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.ChatType.PRIVATE, handle_text))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_input_handler), group=1)
    app.add_handler(MessageHandler((filters.PHOTO | filters.VIDEO | filters.VOICE) & filters.ChatType.PRIVATE, handle_media))

    # Callback handlers
    app.add_handler(CallbackQueryHandler(cancel_input, pattern="^cancel_input$"))
    app.add_handler(CallbackQueryHandler(vehicle_selection_callback, pattern="^selv_"))
    app.add_handler(CallbackQueryHandler(confirm_vehicle_callback, pattern="^confirmveh_"))
    app.add_handler(CallbackQueryHandler(cancel_vehicle_selection_callback, pattern="^cancel_veh$"))
    app.add_handler(CallbackQueryHandler(valide_callback, pattern="^val_"))
    app.add_handler(CallbackQueryHandler(ruglee_callback, pattern="^rug_"))
    app.add_handler(CallbackQueryHandler(fix_comment_callback, pattern="^fix_comment_"))
    app.add_handler(CallbackQueryHandler(comment_callback, pattern="^com_"))
    app.add_handler(CallbackQueryHandler(delete_callback, pattern="^del_"))
    app.add_handler(CallbackQueryHandler(confirm_delete_callback, pattern="^confirmdel_"))
    app.add_handler(CallbackQueryHandler(cancel_delete_callback, pattern="^cancel_"))
    app.add_handler(CallbackQueryHandler(validation_request_callback, pattern="^valreq_"))
    app.add_handler(CallbackQueryHandler(valrug_callback, pattern="^valrug_"))
    app.add_handler(CallbackQueryHandler(close_val_callback, pattern="^close_val_"))
    app.add_handler(CallbackQueryHandler(reopen_val_callback, pattern="^reopen_val_"))
    app.add_handler(CallbackQueryHandler(vidange_confirm_callback, pattern="^vidconfirm_"))
    app.add_handler(CallbackQueryHandler(vidange_modify_callback, pattern="^vidmodify_"))
    app.add_handler(CallbackQueryHandler(approve_callback, pattern="^approve_"))
    app.add_handler(CallbackQueryHandler(reject_callback, pattern="^reject_"))
    app.add_handler(CallbackQueryHandler(admin_main_callback, pattern="^admin_main$"))
    app.add_handler(CallbackQueryHandler(admin_vehicles_menu, pattern="^admin_vehicles$"))
    app.add_handler(CallbackQueryHandler(admin_drivers_menu, pattern="^admin_drivers$"))
    app.add_handler(CallbackQueryHandler(admin_vidange_menu, pattern="^admin_vidange_menu$"))
    app.add_handler(CallbackQueryHandler(admin_export_menu, pattern="^admin_export_menu$"))
    app.add_handler(CallbackQueryHandler(admin_approve_list, pattern="^admin_approve_list$"))
    app.add_handler(CallbackQueryHandler(admin_remove_driver_list, pattern="^admin_remove_driver_list$"))
    app.add_handler(CallbackQueryHandler(admin_drivers_list, pattern="^admin_drivers_list$"))
    app.add_handler(CallbackQueryHandler(confirm_remove_driver, pattern="^rmdriver_"))
    app.add_handler(CallbackQueryHandler(confirm_remove_driver_exec, pattern="^confirmrm_"))
    app.add_handler(CallbackQueryHandler(settings_help, pattern="^settings_help$"))
    app.add_handler(CallbackQueryHandler(settings_history_callback, pattern="^settings_history$"))
    app.add_handler(CallbackQueryHandler(settings_change_name, pattern="^settings_change_name$"))
    app.add_handler(CallbackQueryHandler(cancel_name_callback, pattern="^cancel_name$"))
    app.add_handler(CallbackQueryHandler(settings_change_vehicle, pattern="^settings_change_veh$"))  # NEW
    app.add_handler(CallbackQueryHandler(delete_help_video_callback, pattern="^delhelp_"))
    app.add_handler(CallbackQueryHandler(admin_addveh, pattern="^admin_addveh$"))
    app.add_handler(CallbackQueryHandler(admin_remveh, pattern="^admin_remveh$"))
    app.add_handler(CallbackQueryHandler(admin_setvid, pattern="^admin_setvid$"))
    app.add_handler(CallbackQueryHandler(admin_urgentvid, pattern="^admin_urgentvid$"))
    app.add_handler(CallbackQueryHandler(admin_dash, pattern="^admin_dash$"))
    app.add_handler(CallbackQueryHandler(admin_export, pattern="^admin_export$"))
    app.add_handler(CallbackQueryHandler(admin_vid, pattern="^admin_vid$"))
    app.add_handler(CallbackQueryHandler(admin_listveh, pattern="^admin_listveh$"))
    app.add_handler(CallbackQueryHandler(vehicle_history_callback, pattern="^hist_"))
    app.add_handler(CallbackQueryHandler(done_history_callback, pattern="^done_hist$"))

    app.add_error_handler(error_handler)
    await app.initialize()
    schedule_jobs(app)
    await app.start()
    logging.info("Telegram application and JobQueue started successfully")
    await set_webhook(app)

    aio_app = web.Application(middlewares=[api_error_middleware])
    aio_app.router.add_get("/", health)
    aio_app.router.add_post("/telegram", telegram_webhook)
    register_api_routes(aio_app)

    port = int(os.environ.get("PORT", 5000))
    runner = web.AppRunner(aio_app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logging.info(f"Server listening on port {port}")
    await asyncio.Event().wait()

if __name__ == "__main__":
    asyncio.run(main())
